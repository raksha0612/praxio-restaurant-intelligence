"""
app.py - Restaurant Intelligence Engine v2.0
"""
import base64
import logging
import os
import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env", override=True)
except ImportError:
    pass

import streamlit as st

try:
    import pandas as pd
except Exception:
    pd = None

try:
    import plotly.graph_objects as go
except Exception:
    go = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from data_audit import load_from_google_sheets, find_col, load_and_clean_data, _compute_benchmarks
from scoring_engine import (
    compute_dimension_scores,
    get_gap_analysis,
    compute_momentum,
    get_silent_winner_flag,
    get_customer_persona,
    compute_all_ranks,
    identify_silent_winners,
    calculate_silent_winner_opportunity,
    calculate_deal_probability,
)
from report_generator import generate_pdf_report
from restaurant_chat import (
    build_restaurant_context, get_response, get_suggested_questions,
    get_all_questions, parse_followups,
    get_next_best_action, get_similar_questions,
)
from database import (
    init_db,
    get_call_notes, save_call_note, delete_call_note_by_index,
    save_chat_session, load_chat_session, clear_chat_session,
    save_score_history, get_score_history,
    get_all_restaurants_with_notes,
)
from translations import t
from excel_exporter import export_visit_notes_to_excel

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)-8s] %(name)s: %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

try:
    init_db()
except Exception as _db_init_err:
    logger.warning("DB init failed (non-fatal): %s", _db_init_err)


# ── HELPERS ──────────────────────────────────────────────────────────────────

def get_actionable_solutions(res_name, res_data, top_gaps, language="EN"):
    rating  = float(res_data.get("rating_n", 4.6) or 4.6)
    est_pts = min(int(top_gaps[0][1] * 0.6), 999) if top_gaps else 0
    if language.upper() == "DE":
        return [
            {"emoji": "⚡", "title": "Antwortzeit optimieren",      "priority": "HOCH",    "priority_color": "#DC2626",
             "desc": "Durchschn. Antwortzeit auf unter 2 Stunden senken - wichtigster Umsatzhebel",
             "est": f"Est. +{est_pts} Punkte Score-Steigerung"},
            {"emoji": "⭐", "title": "Bewertungskampagne starten",   "priority": "MITTEL",  "priority_color": "#F59E0B",
             "desc": "Ziel: 15 neue Bewertungen dieses Quartal via Post-Visit SMS",
             "est": "Est. +12% Sichtbarkeit"},
            {"emoji": "🔗", "title": "Google Profil aktualisieren", "priority": "NIEDRIG", "priority_color": "#10B981",
             "desc": "Fotos, Menulinks & Buchungs-CTA aktualisieren",
             "est": "Est. +8% CTR"},
            {"emoji": "🤖", "title": "KI-Bewertungsmanagement",     "priority": "HOCH",    "priority_color": "#DC2626",
             "desc": "Personalisierte Antworten automatisieren - 120 EUR/Mo",
             "est": "Est. 3x Antwortrate"},
            {"emoji": "📡", "title": "Stimmungsüberwachung",        "priority": "MITTEL",  "priority_color": "#F59E0B",
             "desc": "Echtzeit-Benachrichtigungen für negative Bewertungen auf allen Plattformen",
             "est": f"{rating:.1f} Sterne Bewertung schützen"},
        ]
    return [
        {"emoji": "⚡", "title": "Optimize Response Time", "priority": "HIGH",   "priority_color": "#DC2626",
         "desc": "Reduce avg reply to under 2 hours - top revenue lever",
         "est": f"Est. +{est_pts} pts score lift"},
        {"emoji": "⭐", "title": "Launch Review Campaign",  "priority": "MEDIUM", "priority_color": "#F59E0B",
         "desc": "Target 15 new reviews this quarter via post-visit SMS",
         "est": "Est. +12% visibility"},
        {"emoji": "🔗", "title": "Update Google Profile",   "priority": "LOW",    "priority_color": "#10B981",
         "desc": "Refresh photos, menu links & booking CTA",
         "est": "Est. +8% CTR"},
        {"emoji": "🤖", "title": "AI Review Management",    "priority": "HIGH",   "priority_color": "#DC2626",
         "desc": "Automate personalized responses at scale - 120 EUR/mo",
         "est": "Est. 3x response rate"},
        {"emoji": "📡", "title": "Sentiment Monitoring",    "priority": "MEDIUM", "priority_color": "#F59E0B",
         "desc": "Real-time alerts for negative reviews across platforms",
         "est": f"Protect {rating:.1f} star rating"},
    ]


def get_rating_split(res_name, df_rest, df_rev):
    try:
        row  = df_rest[df_rest["name"] == res_name].iloc[0]
        slug = row.get("_slug", "")
        if slug and "_slug" in df_rev.columns:
            sub = df_rev[df_rev["_slug"] == slug]
        else:
            sub = pd.DataFrame()
        if len(sub) > 0 and "review_rating" in sub.columns:
            rc = sub["review_rating"].value_counts().sort_index(ascending=False)
            return rc.values.tolist(), rc.index.tolist()
    except Exception:
        pass
    return [40, 30, 15, 10, 5], [5, 4, 3, 2, 1]


def _pipeline_tag(scores_row, has_notes, lang):
    """Return a (label, color) tuple for the pipeline opportunity tag."""
    score = scores_row.get("Composite", 0)
    res_rate = scores_row.get("res_rate", 1)
    rating = scores_row.get("rating_n", 0)
    # Silent winner: high rating, low response
    if rating >= 4.5 and float(res_rate or 1) < 0.30:
        return t("pipeline_silent_owner", lang), "#8B5CF6"
    # No booking detected via digital presence gap
    if scores_row.get("Digital Presence", 100) < 50:
        return t("pipeline_no_booking", lang), "#F59E0B"
    if score >= 75:
        return t("pipeline_growth", lang), "#22C55E"
    if score >= 60:
        return t("pipeline_on_target", lang), "#3B82F6"
    return t("pipeline_gap", lang), "#EF4444"


def _build_pipeline_excel(df_rest, df_rev, df_ranks_all, lang="EN"):
    """Build the master Excel workbook — pipeline sheet + all visit notes."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Sheet 1: Pipeline / Restaurant scores ──
    ws_pipe = wb.active
    ws_pipe.title = "Pipeline" if lang == "EN" else "Vertriebspipeline"

    HDR_FILL = PatternFill("solid", fgColor="0F172A")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
    BORDER = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
    )

    pipe_headers = [
        "Rank", "Restaurant", "City", "Score", "Rating", "Reviews",
        "Response %", "Opportunity Tag", "Contacted", "Notes Count",
    ]
    for ci, h in enumerate(pipe_headers, 1):
        cell = ws_pipe.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    # Build contacted lookup
    all_rids_with_notes = set(get_all_restaurants_with_notes())

    for row_i, (_, rank_row) in enumerate(df_ranks_all.iterrows(), 2):
        name = rank_row["name"]
        try:
            res_row = df_rest[df_rest["name"] == name].iloc[0]
            s = compute_dimension_scores(name, df_rest, df_rev)
        except Exception:
            continue
        rid = name.lower().replace(" ", "_").replace("-", "_")[:40]
        is_contacted = rid in all_rids_with_notes
        notes_list = get_call_notes(rid)

        tag_label, _ = _pipeline_tag({**dict(res_row), **s}, is_contacted, lang)

        district_col = find_col(df_rest, ["district"])
        city_val = str(res_row.get(district_col, "")) if district_col else ""

        row_data = [
            int(rank_row["rank"]),
            name,
            city_val,
            round(float(s.get("Composite", 0)), 1),
            round(float(res_row.get("rating_n", 0)), 1),
            int(res_row.get("rev_count_n", 0)),
            f"{float(res_row.get('res_rate', 0)) * 100:.0f}%",
            tag_label,
            "Yes" if is_contacted else "No",
            len(notes_list),
        ]
        fill = ALT_FILL if row_i % 2 == 0 else None
        for ci, val in enumerate(row_data, 1):
            cell = ws_pipe.cell(row=row_i, column=ci, value=val)
            cell.border = BORDER
            cell.font = Font(size=9, name="Arial")
            if fill:
                cell.fill = fill

    for col in ws_pipe.columns:
        ws_pipe.column_dimensions[col[0].column_letter].width = min(
            max((len(str(c.value or "")) for c in col), default=0) + 3, 40)

    # ── Sheet 2: Visit Notes (matches Vertriebsreporting Excel) ──
    ws_notes = wb.create_sheet("Visit Notes" if lang == "EN" else "Besuchsnotizen")

    notes_headers = [
        "KW", "Restaurant", "Visit Date", "Time", "City", "District",
        "Price Class", "Size", "Contact Person", "Atmosphere", "Duration",
        "Pre-Check Needs", "Potential (1-10)", "Interest (1-5)",
        "Products Shown", "Outcome", "Next Steps / Follow-up",
        "Detailed Notes", "Self-Reflection (for Kevin)",
        "Main Objection", "Budget", "Confidence %",
        "Decision Timeline", "Competitor Tools",
    ]
    if lang == "DE":
        notes_headers = [
            "KW", "Kunde", "Besuchsdatum", "Uhrzeit", "Stadt", "Stadtteil",
            "Preisklasse", "Größe", "Gesprächspartner", "Stimmung vor Ort", "Dauer des Gesprächs",
            "Bedarf Pre-Check", "Einschätzung Potenzial (1-10)", "Interessensstufe (1-5)",
            "Gezeigte Produkte", "Ergebnis", "Geplantes Follow up",
            "Ausführliche Notizen", "Selbstreflexion (Basis für Dialog mit Kevin)",
            "Haupteinwand", "Budget", "Abschluss-Sicherheit %",
            "Entscheidungsfrist", "Konkurrenztools",
        ]

    for ci, h in enumerate(notes_headers, 1):
        cell = ws_notes.cell(row=1, column=ci, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_idx = 2
    all_rids = get_all_restaurants_with_notes()
    for rid in sorted(all_rids):
        notes_list = get_call_notes(rid)
        res_display = rid.replace("_", " ").title()
        for cn in notes_list:
            prods = ", ".join(cn.get("products_discussed", []) or [])
            # Compute calendar week from visit_date
            kw = ""
            vd = cn.get("visit_date", "")
            if vd:
                try:
                    kw = f"KW {datetime.strptime(str(vd)[:10], '%Y-%m-%d').isocalendar()[1]}"
                except Exception:
                    pass
            ws_notes.append([
                kw,
                res_display,
                cn.get("visit_date", ""),
                cn.get("visit_time", ""),
                cn.get("city", ""),
                cn.get("district", ""),
                cn.get("price_class", ""),
                cn.get("size", ""),
                cn.get("contact_name", ""),
                cn.get("atmosphere", ""),
                cn.get("visit_duration", ""),
                cn.get("pre_check_needs", ""),
                cn.get("potential_score", ""),
                cn.get("interest_level", ""),
                prods,
                cn.get("visit_outcome", cn.get("outcome", "")),
                cn.get("next_steps", ""),
                cn.get("notes", ""),
                cn.get("self_reflection", ""),
                cn.get("main_objection", ""),
                cn.get("budget_range", ""),
                cn.get("confidence", ""),
                cn.get("decision_timeline", ""),
                cn.get("competitor_tools", ""),
            ])
            # Style data rows
            for ci in range(1, len(notes_headers) + 1):
                cell = ws_notes.cell(row=row_idx, column=ci)
                cell.font = Font(size=9, name="Arial")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if row_idx % 2 == 0:
                    cell.fill = ALT_FILL
            row_idx += 1

    # Auto column widths for notes sheet
    for col in ws_notes.columns:
        max_w = max((len(str(c.value or "")) for c in col), default=0)
        ws_notes.column_dimensions[col[0].column_letter].width = min(max_w + 3, 50)
    # Notes columns get more width
    for wide_col in ["R", "S", "O"]:
        ws_notes.column_dimensions[wide_col].width = 50
    # Freeze header row
    ws_notes.freeze_panes = "A2"
    ws_pipe.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── APP CONFIG ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Restaurant Intelligence Engine v2.0",
    page_icon="🍽️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── AUTH ─────────────────────────────────────────────────────────────────────
_APP_PASSWORD = os.environ.get("APP_PASSWORD", "")
if _APP_PASSWORD and not st.session_state.get("authenticated"):
    st.title("PraxioTech · Restaurant Intelligence")
    pwd = st.text_input("Password", type="password")
    if st.button("Login"):
        if pwd == _APP_PASSWORD:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password")
    st.stop()

NAVY, TEAL, TEAL2, BG, WHITE, MUTED, SUCCESS, WARNING, DANGER, BORDER = (
    "#0F172A", "#0EA5E9", "#14B8A6", "#F0F4F8", "#FFFFFF", "#64748B",
    "#22C55E", "#F59E0B", "#EF4444", "#E2E8F0"
)
ACCENT = "#2563EB"
SECONDARY = "#8B5CF6"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=Space+Mono:wght@400;700&display=swap');
html, body, [class*="css"], .stApp {{ font-family: 'Inter', sans-serif !important; }}
#MainMenu, footer, .stDeployButton, [data-testid="stHeader"], [data-testid="stToolbar"] {{ visibility: hidden; display: none; }}
.stApp {{ background: {BG}; }}
.main .block-container {{ padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1400px; }}
@keyframes fadeInSlide {{ from {{ opacity: 0; transform: translateY(12px); }} to {{ opacity: 1; transform: translateY(0); }} }}
@keyframes glow {{ 0%, 100% {{ box-shadow: 0 0 10px rgba(14,165,233,0.5); }} 50% {{ box-shadow: 0 0 20px rgba(14,165,233,0.8); }} }}
label {{ color: {NAVY} !important; font-weight: 600 !important; font-size: 12px !important; }}
[data-testid="stLabel"] {{ color: {NAVY} !important; }}
[data-testid="stLabel"] p {{ color: {NAVY} !important; }}
.stForm label {{ color: {NAVY} !important; }}
.stDateInput label, .stTextInput label, .stSelectbox label, .stSlider label, .stTextArea label, .stMultiSelect label {{ color: {NAVY} !important; }}
input[type="text"], input[type="password"], input[type="number"], input[type="date"], textarea {{
    color: {NAVY} !important; background-color: {WHITE} !important;
    font-size: 12px !important; font-family: 'Inter', sans-serif !important;
}}
input::placeholder, textarea::placeholder {{ color: #94A3B8 !important; font-size: 12px !important; }}
[data-baseweb="input"] input, [data-baseweb="textarea"] textarea {{ color: {NAVY} !important; background-color: {WHITE} !important; font-size: 12px !important; }}
[data-baseweb="input"], [data-baseweb="base-input"], [data-baseweb="textarea"] {{ background-color: {WHITE} !important; }}
.main [data-baseweb="select"] > div {{ background-color: {WHITE} !important; border: 1px solid {BORDER} !important; border-radius: 8px !important; }}
.main [data-baseweb="select"] span, .main [data-baseweb="select"] div,
.main [data-baseweb="select"] input, .main [data-baseweb="select"] p {{ color: {NAVY} !important; background-color: transparent !important; }}
[data-baseweb="menu"], [data-baseweb="popover"] ul, [data-baseweb="popover"] li {{ background-color: {WHITE} !important; color: {NAVY} !important; }}
[data-baseweb="menu"] li:hover, [data-baseweb="menu"] [aria-selected="true"] {{ background-color: #EFF6FF !important; color: {NAVY} !important; }}
.main [data-baseweb="tag"] {{ background-color: #E0F2FE !important; color: {NAVY} !important; }}
[data-testid="stMetricValue"] {{ color: {NAVY} !important; font-size: 13px !important; }}
[data-testid="stMetricLabel"] {{ color: {MUTED} !important; font-size: 10px !important; }}
[data-testid="stSidebar"] > div:first-child {{ background: linear-gradient(175deg, {NAVY} 0%, #101d33 100%) !important; border-right: none !important; }}
[data-testid="stSidebar"] * {{ color: #CBD5E1 !important; }}
[data-testid="stSidebar"] [data-baseweb="select"] > div {{ background: rgba(255,255,255,0.08) !important; border: 1px solid rgba(255,255,255,0.18) !important; border-radius: 8px !important; }}
.kpi-card {{ background: {WHITE}; border-radius: 12px; padding: 18px 22px; box-shadow: 0 1px 3px rgba(0,0,0,0.07); border: 1px solid {BORDER}; position: relative; overflow: hidden; animation: fadeInSlide 0.5s ease-out; }}
.kpi-card::before {{ content: ''; position: absolute; top:0; left:0; right:0; height:3px; background: linear-gradient(90deg, {TEAL}, {TEAL2}); border-radius: 12px 12px 0 0; }}
.kpi-label {{ font-size:10px; font-weight:700; letter-spacing:0.1em; text-transform:uppercase; color:{MUTED}; margin-bottom:6px; }}
.kpi-value {{ font-size:32px; font-weight:800; color:{NAVY}; line-height:1; font-family:'Space Mono',monospace; }}
.kpi-sub {{ font-size:12px; color:{MUTED}; margin-top:8px; }}
.section-card {{ background: {WHITE}; border-radius: 12px; padding: 22px 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.07); border: 1px solid {BORDER}; margin-bottom: 18px; animation: fadeInSlide 0.6s ease-out; }}
.card-header {{ font-size:13px; font-weight:700; color:{NAVY}; padding-bottom:12px; margin-bottom:14px; border-bottom:1px solid {BORDER}; }}
.chat-msg-user {{ background: linear-gradient(135deg, {TEAL}, {TEAL2}); color: white; padding: 14px 18px; border-radius: 20px 20px 4px 20px; margin-left: 15%; margin-bottom: 10px; font-size: 13px; line-height: 1.6; }}
.chat-msg-ai {{ background: {WHITE}; color: {NAVY}; border: 2px solid #CAE6FD; padding: 15px 18px; border-radius: 20px 20px 20px 4px; margin-right: 10%; margin-bottom: 10px; font-size: 13px; line-height: 1.7; }}
[data-testid="stChatMessage"] * {{ color: {NAVY} !important; }}
.stButton > button {{ background: linear-gradient(135deg, {TEAL}, {TEAL2}) !important; color: white !important; border: none !important; border-radius: 8px !important; font-weight: 600 !important; transition: all 0.2s ease !important; }}
.stButton > button:hover {{ transform: translateY(-2px) !important; box-shadow: 0 8px 16px rgba(14,165,233,0.3) !important; }}
[data-testid="stExpander"] summary p, [data-testid="stExpander"] summary {{ color: {NAVY} !important; font-size: 12px !important; font-weight: 600 !important; }}
[data-testid="stExpander"] {{ background: {WHITE} !important; border: 1px solid {BORDER} !important; border-radius: 8px !important; margin-bottom: 4px !important; }}
.pipeline-row {{ background: {WHITE}; border: 1px solid {BORDER}; border-radius: 10px; padding: 12px 16px; margin-bottom: 8px; display: flex; align-items: center; gap: 12px; transition: box-shadow 0.2s; }}
.pipeline-row:hover {{ box-shadow: 0 4px 12px rgba(14,165,233,0.15); }}
</style>
""", unsafe_allow_html=True)

# SESSION STATE
for key, default in [
    ("data_loaded", False), ("df_rest", None), ("df_rev", None), ("benchmarks", None),
    ("chat_messages", []), ("chat_context", None), ("active_page", "dashboard"),
    ("selected_city", "All Cities"), ("language", "EN"),
    ("min_rating_filter", 0.0), ("min_reviews_filter", 0), ("min_response_filter", 0),
]:
    if key not in st.session_state:
        st.session_state[key] = default


@st.cache_data(show_spinner=False)
def load_google_data():
    CITIES = {
        "Frankfurt": {
            "rest": "https://docs.google.com/spreadsheets/d/1GzZWRuPr4y3yscDZYprWZtdgZ6Z6MkPBHwRTLH4bPR8/edit?usp=sharing",
            "rev":  "https://docs.google.com/spreadsheets/d/1zSAd91SkuYgXuIOQa5WVJneEV5dmPyM9XMnI66iNGb4/edit?usp=sharing",
        },
        "Hamburg": {
            "rest": "https://docs.google.com/spreadsheets/d/1W1NmB1wAYh4qJ_tsx6VC2NLthTLSFtliGJYZdK9EaOA/edit?usp=drive_link",
            "rev":  "https://docs.google.com/spreadsheets/d/1HU7hEmKBTn5u42MPDEawdrNDxp7sZZXN13lxACDzSd8/edit?usp=drive_link",
        },
        "Wedel": {
            "rest": "https://docs.google.com/spreadsheets/d/1iOZiX0xy9tZqs5SEOC_7BN5OhKsMMeCdsHJKDZgOQeI/edit?usp=drive_link",
            "rev":  "https://docs.google.com/spreadsheets/d/15MeEt3uZsBLyCqIHqwkzomOXNS22WyaTA2OIU53i89I/edit?usp=drive_link",
        },
    }
    all_rest, all_rev = [], []
    for city_name, urls in CITIES.items():
        try:
            df_r, df_v, _ = load_from_google_sheets(urls["rest"], urls["rev"])
            df_r["district"] = city_name
            all_rest.append(df_r)
            all_rev.append(df_v)
        except Exception as e:
            logger.error(f"Error loading {city_name}: {e}")
    if not all_rest:
        return None, None, None
    df_rest = pd.concat(all_rest, ignore_index=True)
    df_rev  = pd.concat(all_rev,  ignore_index=True)
    benchmarks = _compute_benchmarks(df_rest)
    return df_rest, df_rev, benchmarks


if not st.session_state.data_loaded:
    df_rest, df_rev, benchmarks = load_google_data()
    if df_rest is not None:
        st.session_state.df_rest     = df_rest
        st.session_state.df_rev      = df_rev
        st.session_state.benchmarks  = benchmarks
        st.session_state.data_loaded = True

# ── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style="padding:18px 0 20px;border-bottom:1px solid rgba(255,255,255,0.1);margin-bottom:18px">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
        <span style="font-size:20px">🍽️</span>
        <span style="font-size:17px;font-weight:800;color:white">{t('sidebar_title', st.session_state.language)}</span>
      </div>
      <div style="font-size:11px;color:#475569">{t('sidebar_subtitle', st.session_state.language)}</div>
    </div>""", unsafe_allow_html=True)

    if st.session_state.data_loaded and st.session_state.df_rest is not None:
        _df = st.session_state.df_rest

        st.markdown('<p style="color:#94A3B8;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:6px">🌍 City Filter</p>', unsafe_allow_html=True)
        cities = sorted(_df["district"].dropna().unique().tolist()) if "district" in _df.columns else []
        if "All Cities" not in cities:
            cities = ["All Cities"] + cities
        selected_city = st.selectbox("", cities, label_visibility="collapsed", key="city_select")
        st.session_state.selected_city = selected_city

        if selected_city != "All Cities" and "district" in _df.columns:
            _city_df = _df[_df["district"] == selected_city]
        else:
            _city_df = _df

        st.markdown('<p style="color:#94A3B8;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;margin-top:14px;margin-bottom:6px">⚙️ Advanced Filters</p>', unsafe_allow_html=True)
        _max_rev = int(_df["rev_count_n"].max()) if len(_df) > 0 else 1000
        _min_rat = st.slider(t("min_rating",   st.session_state.language), 0.0, 5.0,      float(st.session_state.min_rating_filter),  0.1, key="filter_rating")
        _min_rev = st.slider(t("min_reviews",  st.session_state.language), 0,   _max_rev, int(st.session_state.min_reviews_filter),    10,  key="filter_reviews")
        _min_res = st.slider(t("min_response", st.session_state.language), 0,   100,      int(st.session_state.min_response_filter),   5,   key="filter_response")
        st.session_state.min_rating_filter   = _min_rat
        st.session_state.min_reviews_filter  = _min_rev
        st.session_state.min_response_filter = _min_res

        _adv_df = _city_df[
            (_city_df["rating_n"] >= _min_rat) &
            (_city_df["rev_count_n"] >= _min_rev) &
            (_city_df["res_rate"].fillna(0) * 100 >= _min_res)
        ]
        if len(_adv_df) == 0:
            _adv_df = _city_df

        st.markdown('<p style="color:#94A3B8;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;margin-top:14px;margin-bottom:6px">🏪 Select Restaurant</p>', unsafe_allow_html=True)
        sidebar_names = sorted(_adv_df["name"].dropna().unique().tolist())
        _prev_sel     = st.session_state.get("rest_select")
        _default_idx  = sidebar_names.index(_prev_sel) if _prev_sel in sidebar_names else 0
        st.selectbox("", sidebar_names, index=_default_idx, label_visibility="collapsed", key="rest_select")

        st.markdown("<hr style='border-color:rgba(255,255,255,0.1);margin:14px 0'>", unsafe_allow_html=True)
        silent_winners_all = identify_silent_winners(_df)
        if silent_winners_all:
            st.markdown(f'<p style="color:#FCA5A5;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:8px">{t("silent_winners_detected", st.session_state.language).format(len(silent_winners_all))}</p>', unsafe_allow_html=True)
        st.markdown(f'<p style="color:#475569;font-size:11px;margin-top:14px">Total: <b>{len(sidebar_names)}</b> restaurants in {selected_city}</p>', unsafe_allow_html=True)

# NO DATA STATE
if not st.session_state.data_loaded:
    st.markdown(f"""<div style="text-align:center;padding:60px 40px">
    <div style="font-size:48px;margin-bottom:20px">🍽️</div>
    <h1 style="font-size:28px;font-weight:800;color:{NAVY};margin-bottom:12px">{t("loading", st.session_state.language)}</h1>
    </div>""", unsafe_allow_html=True)
    st.stop()

# LANGUAGE SELECTOR
col_empty, col_lang = st.columns([4, 1])
with col_lang:
    lang_options  = {"🇬🇧 English": "EN", "🇩🇪 Deutsch": "DE"}
    selected_lang = st.selectbox(t("language", st.session_state.language), list(lang_options.keys()),
                                 index=0 if st.session_state.language == "EN" else 1,
                                 label_visibility="collapsed", key="lang_select")
    if lang_options[selected_lang] != st.session_state.language:
        st.session_state.language = lang_options[selected_lang]
        st.rerun()

lang = st.session_state.language
df_rest, df_rev, benchmarks = st.session_state.df_rest, st.session_state.df_rev, st.session_state.benchmarks

_city = st.session_state.get("selected_city", "All Cities")
if _city and _city != "All Cities" and "district" in df_rest.columns:
    _df_city = df_rest[df_rest["district"] == _city].copy()
else:
    _df_city = df_rest.copy()

df_rest_filtered = _df_city[
    (_df_city["rating_n"] >= st.session_state.min_rating_filter) &
    (_df_city["rev_count_n"] >= st.session_state.min_reviews_filter) &
    (_df_city["res_rate"].fillna(0) * 100 >= st.session_state.min_response_filter)
].copy()
if len(df_rest_filtered) == 0:
    df_rest_filtered = _df_city.copy()

names    = sorted(df_rest_filtered["name"].dropna().unique().tolist())
selected = st.session_state.get("rest_select", names[0] if names else None)

if not selected:
    st.warning(t("no_restaurants", lang))
    st.stop()

# COMPUTE METRICS
res_data    = df_rest[df_rest["name"] == selected].iloc[0]
scores      = compute_dimension_scores(selected, df_rest, df_rev)
gaps        = get_gap_analysis(scores, benchmarks)
momentum    = compute_momentum(selected, df_rev, df_rest)
persona     = get_customer_persona(selected, df_rest, df_rev)
silent_flag = get_silent_winner_flag(selected, df_rest)
deal_prob   = calculate_deal_probability(selected, res_data, scores, gaps)

restaurant_id = selected.lower().replace(" ", "_").replace("-", "_")[:40]

try:
    save_score_history(restaurant_id, {
        "composite":      scores.get("Composite", 0),
        "reputation":     scores.get("Reputation", 0),
        "responsiveness": scores.get("Responsiveness", 0),
        "digital":        scores.get("Digital Presence", 0),
        "visibility":     scores.get("Visibility", 0),
        "intelligence":   scores.get("Intelligence", 0),
    })
except Exception as _sh_err:
    logger.debug("save_score_history non-fatal: %s", _sh_err)

df_ranks_all = compute_all_ranks(df_rest, df_rev)
cur_rank     = int(df_ranks_all[df_ranks_all["name"] == selected]["rank"].values[0])
total        = len(df_ranks_all)

try:
    df_ranks = compute_all_ranks(df_rest_filtered, df_rev) if len(df_rest_filtered) > 0 else pd.DataFrame(columns=["name", "score", "rank"])
except Exception as _e:
    df_ranks = pd.DataFrame(columns=["name", "score", "rank"])

# PAGE HEADER
avg_score = float(benchmarks.get("avg_rating", 4.0)) * 20
is_above_avg = scores["Composite"] >= avg_score
badge_label = t("tag_above_avg", lang) if is_above_avg else t("tag_below_avg", lang)
badge_color = "#22C55E" if is_above_avg else "#F59E0B"

silent_badge = (
    f'&nbsp;&nbsp;<span style="background:#FEE2E2;color:#991B1B;padding:3px 10px;border-radius:20px;'
    f'font-size:11px;font-weight:700">'
    f'{t("silent_winner_badge", lang)}</span>'
) if silent_flag else ""

# City from district column (stamped during multi-city load)
_district_col = find_col(df_rest, ["district"])
_city_display = str(res_data.get(_district_col, "")) if _district_col else ""
_city_prefix = f'<span style="background:#E0F2FE;color:#0369A1;padding:1px 8px;border-radius:10px;font-size:10px;font-weight:700;margin-right:6px">📍 {_city_display}</span>' if _city_display else ""

# Flatten HTML — no line may have 4+ leading spaces inside st.markdown (markdown code-block rule)
st.markdown(
    f'<div style="background:{WHITE};border-radius:12px;padding:18px 26px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,0.07);border:1px solid {BORDER}">'
    f'<div style="display:flex;justify-content:space-between;align-items:center">'
    f'<div>'
    f'<h1 style="font-size:26px;font-weight:800;color:{NAVY};margin:0">{selected}</h1>'
    f'<div style="font-size:12px;color:{MUTED};margin-top:6px">'
    f'{_city_prefix}'
    f'{t("ranked", lang)} <strong>#{cur_rank}</strong> {t("of", lang)} {total} &nbsp;·&nbsp;'
    f'{t("score", lang)}: <strong>{scores["Composite"]:.1f}/100</strong>'
    f'&nbsp;&nbsp;<span style="background:{badge_color}22;color:{badge_color};border:1px solid {badge_color}55;padding:2px 10px;border-radius:20px;font-size:10px;font-weight:700">{badge_label}</span>'
    f'{silent_badge}'
    f'</div>'
    f'</div>'
    f'<div style="text-align:right;font-size:11px;color:{MUTED}">'
    f'<strong>{t("rating", lang)}:</strong> {float(res_data.get("rating_n", 0)):.1f}★ &nbsp;|&nbsp;'
    f'<strong>{t("reviews", lang)}:</strong> {int(res_data.get("rev_count_n", 0)):,}'
    f'</div>'
    f'</div>'
    f'</div>',
    unsafe_allow_html=True
)

# NAV BUTTONS — 5 tabs
col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    if st.button(t("btn_dashboard",      lang), use_container_width=True, key="btn_dashboard"):      st.session_state.active_page = "dashboard"
with col2:
    if st.button(t("btn_assistant",      lang), use_container_width=True, key="btn_assistant"):      st.session_state.active_page = "assistant"
with col3:
    if st.button(t("btn_notes",          lang), use_container_width=True, key="btn_notes"):          st.session_state.active_page = "notes"
with col4:
    if st.button(t("btn_pipeline",       lang), use_container_width=True, key="btn_pipeline"):       st.session_state.active_page = "pipeline"
with col5:
    if st.button(t("btn_silent_winners", lang), use_container_width=True, key="btn_silent_winners"): st.session_state.active_page = "silent_winners"

st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

# ============================================================
# DASHBOARD
# ============================================================
if st.session_state.active_page == "dashboard":
    c1, c2, c3, c4, c5 = st.columns(5)
    tiles = [
        (c1, t("kpi_score",     lang), f"{scores['Composite']:.1f}",                    t("out_of_100",      lang), "⚡ Active", True),
        (c2, t("kpi_rank",      lang), f"#{cur_rank}",                                   f"{total} {t('total', lang, count=total)}", "↗ +2", True),
        (c3, t("kpi_response",  lang), f"{float(res_data.get('res_rate',0))*100:.0f}%",  t("owner_replies",   lang), (f"↗ {t('good', lang)}" if float(res_data.get('res_rate',0))>=0.5 else f"↘ {t('low', lang)}"), float(res_data.get('res_rate',0))>=0.5),
        (c4, t("kpi_sentiment", lang), f"{scores['Intelligence']:.0f}%",                 t("review_sentiment",lang), (f"↗ {t('strong', lang)}" if scores['Intelligence']>=75 else f"↘ {t('needs_work', lang)}"), scores['Intelligence']>=75),
        (c5, t("kpi_freshness", lang), f"{scores['Visibility']:.0f}%",                   t("review_velocity", lang), (f"↗ {t('active', lang)}" if scores['Visibility']>=50 else f"↘ {t('slow_status', lang)}"), scores['Visibility']>=50),
    ]
    for col, label, val, sub, delta, pos in tiles:
        dc = "✅" if pos else "⚠️"
        col.markdown(f"""<div class="kpi-card"><div class="kpi-label">{label}</div><div class="kpi-value">{val}</div><div class="kpi-sub">{sub}&nbsp;<span style="color:{'#22C55E' if pos else '#EF4444'};font-weight:700;">{dc} {delta}</span></div></div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

    col_r, col_g = st.columns(2)
    with col_r:
        st.markdown(f'<div class="section-card"><div class="card-header"> {t("dimension_radar", lang)}</div>', unsafe_allow_html=True)
        dim_labels = ["Reputation","Responsiveness","Digital\nPresence","Intelligence","Visibility"]
        if go is None:
            st.info("Plotly not installed.")
        else:
            sv = [scores[d] for d in ["Reputation","Responsiveness","Digital Presence","Intelligence","Visibility"]]
            bv = [benchmarks.get("rating",4.4)*20, 90, 85, 75, 70]
            fig = go.Figure()
            fig.add_trace(go.Scatterpolar(r=sv+[sv[0]], theta=dim_labels+[dim_labels[0]], fill="toself", name="Score",
                line=dict(color=TEAL, width=2.5), fillcolor="rgba(14,165,233,0.13)", marker=dict(size=5, color=TEAL)))
            fig.add_trace(go.Scatterpolar(r=bv+[bv[0]], theta=dim_labels+[dim_labels[0]], fill="toself", name="Benchmark",
                line=dict(color="#A855F7", width=1.5, dash="dot"), fillcolor="rgba(168,85,247,0.05)", marker=dict(size=4, color="#A855F7")))
            fig.update_layout(polar=dict(bgcolor="white",
                radialaxis=dict(visible=True, range=[0,100], tickfont=dict(size=9, color=NAVY), gridcolor="#F0E2E2"),
                angularaxis=dict(tickfont=dict(color=NAVY, size=11, family="Inter"))),
                legend=dict(orientation="h", yanchor="bottom", y=-0.18, xanchor="center", x=0.5, font=dict(size=11)),
                height=340, margin=dict(l=40,r=40,t=20,b=40), paper_bgcolor="white", plot_bgcolor="white")
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
        st.markdown("</div>", unsafe_allow_html=True)

    with col_g:
        st.markdown(f'<div class="section-card"><div class="card-header"> {t("gap_analysis", lang)}</div>', unsafe_allow_html=True)
        for key, score_key, target in [
            ("responsiveness","Responsiveness",90),("market_sentiment","Intelligence",75),
            ("review_freshness","Visibility",70),("brand_visibility","Digital Presence",80),
            ("reputation","Reputation",benchmarks.get("rating",4.4)*20),
        ]:
            label   = t(key, lang)
            current = scores[score_key]
            diff    = round(current - target, 1)
            color   = SUCCESS if diff >= 0 else DANGER
            sign    = "+" if diff >= 0 else ""
            st.markdown(f"""<div style="margin-bottom:14px"><div style="display:flex;justify-content:space-between;margin-bottom:5px">
              <span style="font-size:12px;font-weight:600;color:{NAVY}">{label}</span><span style="font-size:12px;font-weight:700;color:{color}">{sign}{diff}%</span></div>
              <div style="position:relative;height:8px;background:#F1F5F9;border-radius:4px">
                <div style="position:absolute;left:0;top:0;height:100%;width:{min(current,100)}%;background:linear-gradient(90deg,{TEAL},{TEAL2});border-radius:4px"></div>
                <div style="position:absolute;top:-3px;left:{min(target,100)}%;width:2px;height:14px;background:{NAVY};border-radius:2px"></div>
              </div></div>""", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("---")

    col_ai, col_act = st.columns(2)
    with col_ai:
        st.markdown(f'<div class="section-card"><div class="card-header">{t("customer_insights", lang)}</div>', unsafe_allow_html=True)
        p = persona
        st.markdown(f"""<div style="margin-bottom:12px"><div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#94A3B8;margin-bottom:3px">{t("primary_persona", lang)}</div>
          <div style="font-size:15px;font-weight:700;color:{NAVY}">{p["primary"]}</div></div>
          <div style="margin-bottom:12px"><div style="font-size:10px;font-weight:700;color:#94A3B8">{t("segment", lang)}</div><div style="font-size:13px;color:#334155">{p["segment"]}</div></div>
          <div style="background:#F0F9FF;border:1px solid #BAE6FD;border-radius:8px;padding:10px">
            <span style="font-size:11px;font-weight:700;color:#0369A1">{t("motivation", lang)}</span><span style="font-size:11px;color:{NAVY}"> {p["motivation"]}</span></div>
        </div>""", unsafe_allow_html=True)
        try:
            pdf_bytes = generate_pdf_report(selected, res_data, scores, gaps, momentum, persona, benchmarks, df_rest_filtered, df_rev, cur_rank, total, lang)
            st.download_button(t("export_pdf", lang), data=pdf_bytes, file_name=f"Intelligence_{selected.replace(' ','_')}.pdf", mime="application/pdf")
        except Exception as e:
            st.error(f"PDF error: {e}")
        st.markdown("</div>", unsafe_allow_html=True)

    with col_act:
        st.markdown(f'<div class="section-card"><div class="card-header"> {t("actionable_solutions", lang)}</div>', unsafe_allow_html=True)
        top_gaps  = sorted(gaps.items(), key=lambda x: x[1], reverse=True)[:5]
        solutions = get_actionable_solutions(selected, res_data, top_gaps, lang)
        for sol in solutions:
            st.markdown(f"""
            <div style="background:linear-gradient(135deg,{sol['priority_color']}15,{sol['priority_color']}05);border:1px solid {sol['priority_color']}40;border-radius:10px;padding:14px;margin-bottom:10px;border-left:4px solid {sol['priority_color']}">
              <div style="display:flex;align-items:flex-start;gap:12px">
                <div style="font-size:20px">{sol['emoji']}</div>
                <div style="flex:1">
                  <div style="display:flex;align-items:center;gap:8px;margin-bottom:4px">
                    <span style="font-size:13px;font-weight:700;color:{NAVY}">{sol['title']}</span>
                    <span style="background:{sol['priority_color']};color:white;padding:2px 8px;border-radius:4px;font-size:9px;font-weight:700">{sol['priority']}</span>
                  </div>
                  <div style="font-size:11px;color:{MUTED};margin-bottom:6px;line-height:1.5">{sol['desc']}</div>
                  <div style="font-size:10px;font-weight:600;color:{TEAL}">→ {sol['est']}</div>
                </div>
              </div>
            </div>""", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown(f'<div class="section-card"><div class="card-header"> {t("momentum", lang)}</div>', unsafe_allow_html=True)
    col_m1, col_m2, col_m3 = st.columns([3,1,1])
    with col_m1:
        if momentum is not None and len(momentum) > 0 and go is not None:
            fig_m = go.Figure()
            fig_m.add_trace(go.Scatter(x=momentum["month"], y=momentum["count"], mode="lines+markers",
                line=dict(color=TEAL, width=2.5, shape="spline"), fill="tozeroy", fillcolor="rgba(14,165,233,0.08)",
                marker=dict(size=6, color=TEAL, line=dict(color="white", width=1.5))))
            fig_m.update_layout(height=210, margin=dict(l=40,r=20,t=10,b=30), paper_bgcolor="white", plot_bgcolor="white",
                showlegend=False, xaxis=dict(tickfont=dict(color=NAVY)), yaxis=dict(tickfont=dict(color=NAVY)))
            st.plotly_chart(fig_m, use_container_width=True, config={"displayModeBar": False})
    with col_m2:
        if go is not None:
            rc_values, rc_index = get_rating_split(selected, df_rest, df_rev)
            fig_d = go.Figure(go.Pie(values=rc_values, labels=[f"{i}★" for i in rc_index], hole=0.6,
                marker=dict(colors=["#22C55E","#86EFAC","#FCD34D","#FCA5A5","#EF4444"]), textfont=dict(size=9, color=NAVY)))
            fig_d.update_layout(title=dict(text=t("rating_split", lang), font=dict(size=11), x=0.5),
                height=210, margin=dict(l=0,r=0,t=30,b=0), paper_bgcolor="white")
            st.plotly_chart(fig_d, use_container_width=True, config={"displayModeBar": False})
    with col_m3:
        if go is not None:
            fig_g = go.Figure(go.Indicator(mode="gauge+number", value=scores["Composite"],
                number={"font": {"size": 26, "family": "Space Mono", "color": NAVY}},
                gauge={"axis": {"range": [0,100], "tickfont": {"size":8, "color": NAVY}},
                       "bar": {"color": TEAL, "thickness": 0.28}, "bgcolor": "#F1F5F9", "bordercolor": BORDER,
                       "steps": [{"range":[0,50],"color":"#FEE2E2"},{"range":[50,75],"color":"#FEF3C7"},{"range":[75,100],"color":"#DCFCE7"}]}))
            fig_g.update_layout(title=dict(text=t("health", lang), font=dict(size=11), x=0.5),
                height=210, margin=dict(l=20,r=20,t=30,b=0), paper_bgcolor="white")
            st.plotly_chart(fig_g, use_container_width=True, config={"displayModeBar": False})
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown(f'<div class="section-card"><div class="card-header"> {t("top_restaurants", lang)}</div>', unsafe_allow_html=True)
    top_10 = df_ranks.head(10)
    if go is not None and len(top_10) > 0:
        fig_top = go.Figure(go.Bar(x=top_10["name"], y=top_10["score"],
            marker=dict(color=top_10["score"], colorscale="RdYlGn", cmin=0, cmax=100, showscale=False),
            text=top_10["score"].round(1), textposition="auto"))
        fig_top.update_layout(height=300, margin=dict(l=50,r=20,t=10,b=80), paper_bgcolor="white", plot_bgcolor="white",
            xaxis_tickangle=-45, showlegend=False, font=dict(family="Inter", color=NAVY, size=10),
            xaxis=dict(tickfont=dict(color=NAVY, size=10), showgrid=False),
            yaxis=dict(tickfont=dict(color=NAVY, size=10), showgrid=True, gridcolor="#F0F0F0", gridwidth=1))
        fig_top.update_traces(textposition="outside", textfont=dict(color=NAVY, size=10))
        st.plotly_chart(fig_top, use_container_width=True, config={"displayModeBar": False})
    st.markdown("</div>", unsafe_allow_html=True)

# ============================================================
# AI ASSISTANT
# ============================================================
elif st.session_state.active_page == "assistant":
    st.markdown(f"""
    <div style="background:linear-gradient(135deg, {TEAL}, {TEAL2});border-radius:12px;padding:20px 24px;margin-bottom:20px;box-shadow:0 8px 24px rgba(14,165,233,0.25)">
      <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.15em;color:rgba(255,255,255,0.85);margin-bottom:6px">{t("ai_assistant_header", lang)}</div>
      <div style="font-size:18px;font-weight:800;color:white;margin-bottom:3px">{t("ai_assistant_subtitle", lang)} <span style="background:rgba(255,255,255,0.2);padding:4px 10px;border-radius:6px;margin-left:6px">{selected}</span></div>
      <div style="font-size:12px;color:rgba(255,255,255,0.8);margin-top:6px;line-height:1.5">{t("ai_assistant_description", lang)}</div>
      <div style="font-size:10px;color:rgba(255,255,255,0.7);margin-top:8px"> {t("powered_by", lang)}</div>
    </div>""", unsafe_allow_html=True)

    if st.session_state.chat_context is None or st.session_state.get("_chat_restaurant") != selected:
        st.session_state.chat_context     = build_restaurant_context(selected, res_data, scores, gaps, benchmarks, df_rest, df_rev, cur_rank, total, persona, momentum)
        st.session_state.chat_messages    = load_chat_session(restaurant_id)
        st.session_state._chat_restaurant = selected

    user_input = st.chat_input(t("ask_placeholder", lang, selected=selected))
    if user_input or st.session_state.get("pending_question"):
        question = st.session_state.get("pending_question") or user_input
        st.session_state.pending_question = None
        st.session_state.chat_messages.append({"role": "user", "content": question})
        with st.spinner(t("claude_thinking", lang)):
            response = get_response(st.session_state.chat_messages, st.session_state.chat_context, lang)
        st.session_state.chat_messages.append({"role": "assistant", "content": response})
        try:
            save_chat_session(restaurant_id, _city or "", st.session_state.chat_messages)
        except Exception:
            pass
        st.rerun()

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    if st.session_state.chat_messages:
        for msg in st.session_state.chat_messages:
            if msg["role"] == "user":
                st.chat_message("user").markdown(msg["content"])
            else:
                display_text, _ = parse_followups(msg["content"])
                st.chat_message("assistant").markdown(display_text)

        last_ai = next((m for m in reversed(st.session_state.chat_messages) if m["role"] == "assistant"), None)
        if last_ai:
            _, followups = parse_followups(last_ai["content"])
            if followups:
                st.markdown(f"<div style='margin:20px 0 12px;'><p style='font-size:11px;color:#64748B;font-weight:700;margin-bottom:8px'> {t('followup_questions', lang)}</p>", unsafe_allow_html=True)
                for i, q in enumerate(followups):
                    if st.button(f"💬 {q[:50]}{'...' if len(q)>50 else ''}", key=f"fu_{i}_{hash(q)}", use_container_width=True):
                        st.session_state.pending_question = q
                        st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)
            if len(st.session_state.chat_messages) >= 2:
                last_user = next((m for m in reversed(st.session_state.chat_messages) if m["role"] == "user"), None)
                if last_user and last_ai:
                    with st.spinner(t("generating_questions", lang)):
                        similar = get_similar_questions(last_user["content"], last_ai["content"][:800], selected, gaps, language=lang)
                    if similar:
                        st.markdown(f"<div style='margin:16px 0 12px;'><p style='font-size:11px;color:#64748B;font-weight:700;margin-bottom:8px'> {t('similar_questions', lang)}</p>", unsafe_allow_html=True)
                        for i, q in enumerate(similar):
                            if st.button(f"🔍 {q[:50]}{'...' if len(q)>50 else ''}", key=f"sim_{i}_{hash(q)}", use_container_width=True):
                                st.session_state.pending_question = q
                                st.rerun()
                        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button(t("clear_chat", lang), use_container_width=True):
            st.session_state.chat_messages = []
            try:
                clear_chat_session(restaurant_id)
            except Exception:
                pass
            st.rerun()
    else:
        st.markdown(f'<p style="font-size:11px;color:{MUTED};margin-bottom:8px"> {t("suggested_questions", lang)}</p>', unsafe_allow_html=True)
        suggested = get_suggested_questions(gaps, selected, language=lang)
        for i, q in enumerate(suggested[:4]):
            if st.button(q, key=f"sq_{i}", use_container_width=True):
                st.session_state.pending_question = q
                st.rerun()

# ============================================================
# VISIT LOG (Call Notes — updated to match Excel template)
# ============================================================
elif st.session_state.active_page == "notes":
    existing_calls = get_call_notes(restaurant_id)

    st.markdown(f"""
    <div style="background:{WHITE};border-radius:12px;padding:18px 24px;margin-bottom:20px;border:1px solid {BORDER};box-shadow:0 1px 3px rgba(0,0,0,0.07)">
      <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.12em;color:#94A3B8;margin-bottom:4px">{t('call_notes_header', lang)}</div>
      <div style="font-size:18px;font-weight:800;color:{NAVY}">{selected}</div>
      <div style="font-size:12px;color:{MUTED};margin-top:3px">
        {t('calls_logged', lang).format(len(existing_calls))} &nbsp;·&nbsp;
        <span style="color:{TEAL};font-weight:600">{t('active_account', lang)}</span>
      </div>
    </div>""", unsafe_allow_html=True)

    col_exp1, col_exp2, _ = st.columns([1, 1, 2])
    with col_exp1:
        if st.button(t("export_excel", lang), use_container_width=True, key="export_btn"):
            try:
                excel_bytes = export_visit_notes_to_excel(
                    lang=lang,
                    df_rest=df_rest,
                    df_rev=df_rev,
                    df_ranks_all=df_ranks_all,
                    compute_scores_fn=compute_dimension_scores,
                    find_col_fn=find_col,
                )
                st.download_button(
                    "⬇️ Download Excel",
                    data=excel_bytes,
                    file_name=f"Vertriebsreporting_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_excel_btn",
                )
            except Exception as e:
                import traceback
                st.error(f"Export Error: {e}\n{traceback.format_exc()}")
    with col_exp2:
        if st.button(t("refresh_data", lang), use_container_width=True, key="refresh_notes_btn"):
            st.rerun()

    st.markdown(f"""
    <div style="font-size:13px;font-weight:700;padding:12px 16px;background:linear-gradient(135deg,{TEAL},{TEAL2});color:white;border-radius:10px;margin:20px 0 14px;letter-spacing:0.03em">
      {t('log_new_visit', lang)}
    </div>""", unsafe_allow_html=True)

    with st.form("visit_form", clear_on_submit=True):
        # ── Row 1: Visit Details ──
        st.markdown(f'<p style="font-size:11px;font-weight:700;color:{TEAL};margin-bottom:4px;text-transform:uppercase;letter-spacing:0.08em">{t("section_visit_details", lang)}</p>', unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            visit_date = st.date_input(t("visit_date", lang), key="visit_date_input")
        with col2:
            visit_time = st.text_input(t("visit_time", lang), placeholder="14:30", key="visit_time_input")
        with col3:
            visit_city = st.text_input(t("city", lang), value=_city if _city != "All Cities" else "", key="visit_city_input")
        with col4:
            visit_district = st.text_input(t("district", lang), key="visit_district_input")

        col5, col6, col7 = st.columns(3)
        with col5:
            _price_opts = (
                ["", "€  (under €15)", "€€  (€15–30)", "€€€  (€30–60)", "€€€€  (€60+)"]
                if lang == "EN"
                else ["", "€  (unter 15€)", "€€  (15–30€)", "€€€  (30–60€)", "€€€€  (60€+)"]
            )
            price_class = st.selectbox(t("price_class", lang), _price_opts, key="price_class_input")
        with col6:
            size_options = (
                ["", "Klein / Small", "Mittel / Medium", "Groß / Large"]
                if lang == "EN"
                else ["", "Klein", "Mittel", "Groß"]
            )
            size = st.selectbox(t("size", lang), size_options, key="size_input")
        with col7:
            visit_duration = st.text_input(t("visit_duration", lang), placeholder="15 min." if lang == "EN" else "15 Min.", key="duration_input")

        # ── Row 2: Contact & Atmosphere ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        col8, col9 = st.columns(2)
        with col8:
            contact_name = st.text_input(t("contact_name", lang), placeholder=t("contact_placeholder", lang), key="contact_name_input")
        with col9:
            atmosphere = st.text_input(t("atmosphere", lang), placeholder=t("atmosphere_placeholder", lang), key="atmosphere_input")

        # ── Row 3: Pre-Check & Scoring ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        col10, col11, col12 = st.columns([3, 1, 1])
        with col10:
            pre_check = st.text_input(t("pre_check_needs", lang), placeholder=t("pre_check_placeholder", lang), key="precheck_input")
        with col11:
            potential = st.slider(t("potential_score", lang), 1, 10, 5, key="potential_input")
        with col12:
            interest = st.slider(t("interest_level", lang), 1, 5, 3, key="interest_input", help="1=Kalt/Cold  5=Heiß/Hot")

        # ── Row 4: Sales Context ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        st.markdown(f'<p style="font-size:11px;font-weight:700;color:{TEAL};margin-bottom:4px;text-transform:uppercase;letter-spacing:0.08em">{t("section_sales_context", lang)}</p>', unsafe_allow_html=True)
        col13, col14, col15 = st.columns(3)
        with col13:
            objection = st.text_input(t("main_objection", lang), placeholder=t("objection_placeholder", lang), key="objection_input")
        with col14:
            budget = st.text_input(t("budget_range", lang), placeholder=t("budget_placeholder", lang), key="budget_input")
        with col15:
            confidence = st.slider(t("confidence_level", lang), 0, 100, 50, 10, key="confidence_input")

        col16, col17 = st.columns(2)
        with col16:
            decision_timeline = st.text_input(t("decision_timeline", lang), placeholder=t("timeline_placeholder", lang), key="timeline_input")
        with col17:
            competitor_tools = st.text_input(t("competitor_tools", lang), placeholder=t("competitor_placeholder", lang), key="competitor_input")

        # ── Row 5: Products & Outcome ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        col_prod, col_out = st.columns([2, 1])
        with col_prod:
            st.markdown(f'<p style="font-size:11px;font-weight:700;color:{TEAL};margin-bottom:4px;text-transform:uppercase;letter-spacing:0.08em">{t("section_products", lang)}</p>', unsafe_allow_html=True)
            products = st.multiselect(
                t("products_discussed", lang),
                ["AI Review Manager (120 EUR/mo)", "Review Velocity (80 EUR/mo)",
                 "Profile Optimization (60 EUR/mo)", "Sentiment Monitoring (80 EUR/mo)",
                 "Engagement Booster (60 EUR/mo)", "Full Suite (340 EUR/mo)",
                 "Website", "QR Code / Cube", "App Demo", "Visitenkarte"],
                key="products_input", label_visibility="collapsed",
            )
        with col_out:
            st.markdown(f'<p style="font-size:11px;font-weight:700;color:{TEAL};margin-bottom:4px;text-transform:uppercase;letter-spacing:0.08em">{t("section_outcome", lang)}</p>', unsafe_allow_html=True)
            outcome_options = [
                t("outcome_pending", lang), t("outcome_interested", lang),
                t("outcome_demo_scheduled", lang), t("outcome_proposal_sent", lang),
                t("outcome_won", lang), t("outcome_lost", lang),
            ]
            outcome = st.selectbox(t("outcome", lang) if False else "", outcome_options, key="outcome_input", label_visibility="collapsed")

        # ── Row 6: Follow-up ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        st.markdown(f'<p style="font-size:11px;font-weight:700;color:{TEAL};margin-bottom:4px;text-transform:uppercase;letter-spacing:0.08em">{t("section_followup", lang)}</p>', unsafe_allow_html=True)
        col18, col19 = st.columns(2)
        with col18:
            next_steps = st.text_input(t("next_steps", lang), placeholder=t("next_steps_placeholder", lang), key="next_steps_input")
        with col19:
            follow_up_date = st.date_input(t("followup_date", lang), key="followup_date_input")

        # ── Row 7: Notes + Self-Reflection ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        notes = st.text_area(t("detailed_notes", lang), height=100, placeholder=t("notes_placeholder", lang), key="notes_input")
        self_reflection = st.text_area(t("self_reflection", lang), height=80, placeholder=t("reflection_placeholder", lang), key="reflection_input")

        # ── Images ──
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:12px 0'>", unsafe_allow_html=True)
        st.markdown(f'<p style="font-size:11px;font-weight:700;color:{TEAL};margin-bottom:4px;text-transform:uppercase;letter-spacing:0.08em">{t("section_images", lang)}</p>', unsafe_allow_html=True)
        st.markdown(f'<p style="font-size:11px;color:{MUTED};margin-bottom:8px">{t("images_hint", lang)}</p>', unsafe_allow_html=True)
        uploaded_images = st.file_uploader("Upload Images", type=["png","jpg","jpeg","webp"],
                                           accept_multiple_files=True, key="images_input", label_visibility="collapsed")

        submitted = st.form_submit_button(t("save_visit", lang), use_container_width=True)

    if submitted:
        images_b64 = []
        if uploaded_images:
            for img_file in uploaded_images:
                try:
                    b64 = base64.b64encode(img_file.read()).decode("utf-8")
                    images_b64.append({"filename": img_file.name, "type": img_file.type, "data": b64})
                except Exception as img_err:
                    logger.warning(f"Could not encode image {img_file.name}: {img_err}")
        visit_record = {
            "visit_date":        str(visit_date),
            "visit_time":        visit_time,
            "city":              visit_city,
            "district":          visit_district,
            "price_class":       price_class,
            "size":              size,
            "contact_name":      contact_name,
            "atmosphere":        atmosphere,
            "visit_duration":    visit_duration,
            "pre_check_needs":   pre_check,
            "potential_score":   potential,
            "interest_level":    interest,
            "main_objection":    objection,
            "budget_range":      budget,
            "confidence":        confidence,
            "decision_timeline": decision_timeline,
            "competitor_tools":  competitor_tools,
            "products_discussed": products,
            "visit_outcome":     outcome,
            "next_steps":        next_steps,
            "followup_date":     str(follow_up_date),
            "notes":             notes,
            "self_reflection":   self_reflection,
            "image_data":        json.dumps(images_b64),
        }
        save_call_note(restaurant_id, visit_record)
        st.session_state.chat_context = None
        img_msg = f" + {len(images_b64)} image(s)" if images_b64 else ""
        st.success(t("visit_saved", lang).format(selected) + img_msg)
        st.rerun()

    # ── Previous Visits ──
    if existing_calls:
        st.markdown("---")
        st.markdown(f"""
        <div style="font-size:13px;font-weight:700;color:white;background:linear-gradient(135deg,{TEAL},{TEAL2});
                    padding:11px 16px;border-radius:10px;margin-bottom:16px">
          {t('previous_visits', lang).format(len(existing_calls))}
        </div>""", unsafe_allow_html=True)

        for i, call in enumerate(reversed(existing_calls), 1):
            actual_index = len(existing_calls) - i
            interest_val = call.get("interest_level", 0) or 0
            potential_val = call.get("potential_score", "—") or "—"
            outcome_val  = call.get("visit_outcome", call.get("outcome", "Pending")) or "Pending"
            call_images  = call.get("images", [])
            int_stars    = "★" * int(interest_val) + "☆" * (5 - int(interest_val))

            out_colors = {
                t("outcome_won", lang): SUCCESS, t("outcome_lost", lang): DANGER,
                "Won": SUCCESS, "Lost": DANGER, "Gewonnen": SUCCESS, "Verloren": DANGER,
            }
            out_color = out_colors.get(outcome_val, WARNING)

            with st.container():
                c_left, c_right = st.columns([3, 1])
                with c_left:
                    date_str = call.get("visit_date", call.get("call_date", "?"))
                    time_str = call.get("visit_time", "")
                    loc_str  = " · ".join(filter(None, [call.get("city",""), call.get("district","")]))
                    st.markdown(f'<div style="font-size:14px;font-weight:800;color:{NAVY};margin-bottom:2px">{t("visit_number", lang).format(len(existing_calls)-i+1, date_str)} {time_str} <span style="font-weight:400;font-size:11px;color:{MUTED}">{loc_str}</span></div>', unsafe_allow_html=True)
                with c_right:
                    st.markdown(f'<div style="text-align:right"><span style="background:{out_color}22;color:{out_color};border:1px solid {out_color}55;padding:3px 10px;border-radius:20px;font-size:10px;font-weight:700">{outcome_val.upper()}</span></div>', unsafe_allow_html=True)

                m1, m2, m3, m4, m5 = st.columns(5)
                m1.metric(t("contact_name", lang),   call.get("contact_name","—") or "—")
                m2.metric(t("interest_level", lang),  f"{int_stars} {interest_val}/5")
                m3.metric(t("potential_score", lang), f"{potential_val}/10")
                m4.metric(t("budget_range", lang),    call.get("budget_range","—") or "—")
                m5.metric(t("confidence_level", lang), f"{call.get('confidence','—')}%")

                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    prods = ", ".join(call.get("products_discussed", [])) or "—"
                    atm   = call.get("atmosphere", "—") or "—"
                    dur   = call.get("visit_duration", "—") or "—"
                    st.markdown(f'<div style="font-size:11px;color:{NAVY};padding:6px 0"><b>{t("atmosphere", lang)}:</b> {atm} ({dur})<br><b>{t("products_discussed", lang)}:</b> {prods}<br><b>{t("main_objection", lang)}:</b> {call.get("main_objection","—") or "—"}</div>', unsafe_allow_html=True)
                with col_d2:
                    st.markdown(f'<div style="font-size:11px;color:{NAVY};padding:6px 0"><b>{t("next_steps", lang)}:</b> {call.get("next_steps","—") or "—"}<br><b>{t("followup_date", lang)}:</b> {call.get("followup_date","—")}<br><b>{t("pre_check_needs", lang)}:</b> {call.get("pre_check_needs","—") or "—"}</div>', unsafe_allow_html=True)

                if call.get("notes"):
                    st.markdown(f'<div style="font-size:11px;color:{NAVY};background:#F0F9FF;border-left:3px solid {TEAL};padding:9px 12px;border-radius:6px;margin:6px 0 4px;line-height:1.6">📝 {call["notes"]}</div>', unsafe_allow_html=True)
                if call.get("self_reflection"):
                    st.markdown(f'<div style="font-size:11px;color:{NAVY};background:#F0FDF4;border-left:3px solid #22C55E;padding:9px 12px;border-radius:6px;margin:4px 0 8px;line-height:1.6">🧠 {call["self_reflection"]}</div>', unsafe_allow_html=True)

                if call_images:
                    st.markdown(f'<p style="font-size:10px;font-weight:700;color:{TEAL};margin:4px 0 6px;text-transform:uppercase;letter-spacing:0.08em">📸 Attached Images ({len(call_images)})</p>', unsafe_allow_html=True)
                    for ci, img_data in enumerate(call_images):
                        try:
                            img_bytes = base64.b64decode(img_data["data"])
                            fname     = img_data.get("filename", f"Image {ci+1}")
                            st.markdown(f"""<div style="display:inline-flex;align-items:center;gap:8px;background:#F0F9FF;border:1px solid #BAE6FD;border-radius:8px;padding:6px 12px;margin-bottom:4px">
                              <span style="font-size:16px">🖼️</span>
                              <span style="font-size:11px;font-weight:600;color:{NAVY}">{fname}</span></div>""", unsafe_allow_html=True)
                            with st.expander(f"🔍 {fname}"):
                                st.image(img_bytes, use_container_width=True)
                        except Exception:
                            pass

                del_col, _ = st.columns([1, 5])
                with del_col:
                    if st.button(t("delete_visit", lang), key=f"del_{restaurant_id}_{actual_index}", use_container_width=True):
                        ok = delete_call_note_by_index(restaurant_id, actual_index)
                        st.success(t("visit_deleted", lang)) if ok else st.error(t("delete_error", lang))
                        st.rerun()

                st.markdown(f'<hr style="border:none;border-top:1px solid {BORDER};margin:12px 0">', unsafe_allow_html=True)

# ============================================================
# SALES PIPELINE
# ============================================================
elif st.session_state.active_page == "pipeline":
    # Header
    st.markdown(f"""
    <div style="background:{NAVY};border-radius:12px;padding:20px 24px;margin-bottom:20px;color:white">
      <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.15em;color:rgba(255,255,255,0.6);margin-bottom:4px">{t('pipeline_header', lang)}</div>
      <div style="font-size:18px;font-weight:800;color:white;margin-bottom:6px">{t('pipeline_subtitle', lang)} — {len(df_ranks_all)} Restaurants</div>
      <div style="font-size:12px;color:rgba(255,255,255,0.75)">{t('pipeline_desc', lang)}</div>
    </div>""", unsafe_allow_html=True)

    # KPI row
    all_rids_with_notes = set(get_all_restaurants_with_notes())
    total_rest   = len(df_ranks_all)
    contacted    = sum(1 for _, r in df_ranks_all.iterrows()
                       if r["name"].lower().replace(" ","_").replace("-","_")[:40] in all_rids_with_notes)
    uncontacted  = total_rest - contacted
    all_gaps     = []
    for _, r in df_ranks_all.iterrows():
        try:
            s = compute_dimension_scores(r["name"], df_rest, df_rev)
            g = get_gap_analysis(s, benchmarks)
            all_gaps.append(max(g.values()) if g else 0)
        except Exception:
            pass
    avg_gap = round(sum(all_gaps) / len(all_gaps), 1) if all_gaps else 0

    k1, k2, k3, k4 = st.columns(4)
    for col, label, val, sub, color in [
        (k1, t("pipeline_total", lang),        str(total_rest),  "",                         TEAL),
        (k2, t("pipeline_uncontacted", lang),   str(uncontacted), t("pipeline_fresh_leads", lang), "#F59E0B"),
        (k3, t("pipeline_contacted", lang),     str(contacted),   t("pipeline_in_pipeline", lang), "#22C55E"),
        (k4, t("pipeline_avg_gap", lang),        f"{avg_gap} pts","",                         "#EF4444"),
    ]:
        col.markdown(f"""<div class="kpi-card">
          <div class="kpi-label">{label}</div>
          <div class="kpi-value" style="font-size:28px;color:{color}">{val}</div>
          <div class="kpi-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

    # Filters
    col_f1, col_f2, col_f3 = st.columns([2, 2, 2])
    with col_f1:
        all_districts = ["All"] + sorted(df_rest["district"].dropna().unique().tolist()) if "district" in df_rest.columns else ["All"]
        pipe_district = st.selectbox(t("pipeline_filter_district", lang), all_districts, key="pipe_district")
    with col_f2:
        status_opts = [
            t("pipeline_status_all", lang),
            t("pipeline_status_uncontacted", lang),
            t("pipeline_status_contacted", lang),
        ]
        pipe_status = st.selectbox(t("pipeline_filter_status", lang), status_opts, key="pipe_status")
    with col_f3:
        show_top = st.slider(t("pipeline_show_top", lang), 5, min(len(df_ranks_all), 50), min(20, len(df_ranks_all)), key="pipe_top")

    st.markdown(f'<div style="font-size:11px;font-weight:700;text-transform:uppercase;color:#94A3B8;letter-spacing:0.1em;margin:16px 0 10px">OPPORTUNITY RANKING</div>', unsafe_allow_html=True)

    # Build opportunity list
    opp_rows = []
    for _, rank_row in df_ranks_all.iterrows():
        name = rank_row["name"]
        try:
            res_row = df_rest[df_rest["name"] == name].iloc[0]
            s = compute_dimension_scores(name, df_rest, df_rev)
            g = get_gap_analysis(s, benchmarks)
        except Exception:
            continue
        rid = name.lower().replace(" ", "_").replace("-", "_")[:40]
        is_contacted = rid in all_rids_with_notes
        notes_count  = len(get_call_notes(rid))

        district_col = find_col(df_rest, ["district"])
        city_val = str(res_row.get(district_col, "")) if district_col else ""

        # Filter by district
        if pipe_district != "All" and city_val != pipe_district:
            continue
        # Filter by status
        if pipe_status == t("pipeline_status_uncontacted", lang) and is_contacted:
            continue
        if pipe_status == t("pipeline_status_contacted", lang) and not is_contacted:
            continue

        # Opportunity score = biggest gap * weight
        largest_gap = max(g.values()) if g else 0
        opp_score = int(largest_gap * 1.2) if not is_contacted else int(largest_gap * 0.8)
        gap_pts = -round(max(g.values()), 0) if g else 0

        tag_label, tag_color = _pipeline_tag(
            {**dict(res_row), **s, "res_rate": float(res_row.get("res_rate", 1))},
            is_contacted, lang
        )
        opp_rows.append({
            "rank": int(rank_row["rank"]),
            "name": name,
            "city": city_val,
            "score": round(float(s.get("Composite", 0)), 1),
            "rating": round(float(res_row.get("rating_n", 0)), 1),
            "reviews": int(res_row.get("rev_count_n", 0)),
            "res_rate": float(res_row.get("res_rate", 0)),
            "opp_score": opp_score,
            "gap_pts": gap_pts,
            "tag_label": tag_label,
            "tag_color": tag_color,
            "contacted": is_contacted,
            "notes_count": notes_count,
        })

    # Sort by opp_score desc
    opp_rows.sort(key=lambda x: x["opp_score"], reverse=True)

    for idx, row in enumerate(opp_rows[:show_top], 1):
        contacted_dot = f'<span style="color:#22C55E;font-size:10px">●</span>' if row["contacted"] else f'<span style="color:#94A3B8;font-size:10px">○</span>'
        notes_badge   = f'<span style="background:#E0F2FE;color:#0369A1;padding:2px 7px;border-radius:4px;font-size:9px;font-weight:700;margin-left:6px">📋 {row["notes_count"]}</span>' if row["notes_count"] > 0 else ""
        gap_color     = "#EF4444" if row["gap_pts"] < -15 else "#F59E0B" if row["gap_pts"] < -5 else "#22C55E"

        st.markdown(f"""
        <div style="background:{WHITE};border:1px solid {BORDER};border-radius:10px;padding:13px 16px;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 1px 3px rgba(0,0,0,0.05)">
          <div style="display:flex;align-items:center;gap:12px;flex:1">
            <span style="font-size:12px;font-weight:800;color:{MUTED};min-width:26px">#{idx}</span>
            <div>
              <div style="font-size:13px;font-weight:700;color:{NAVY}">{row['name']} {contacted_dot} {notes_badge}</div>
              <div style="font-size:10px;color:{MUTED};margin-top:2px">{row['city']} &nbsp;·&nbsp; {row['score']}/100</div>
            </div>
          </div>
          <div style="display:flex;align-items:center;gap:10px">
            <span style="background:{row['tag_color']}22;color:{row['tag_color']};border:1px solid {row['tag_color']}55;padding:3px 9px;border-radius:20px;font-size:9px;font-weight:700">{row['tag_label']}</span>
            <span style="color:{gap_color};font-size:11px;font-weight:700;min-width:50px;text-align:right">{row['gap_pts']:+.0f} pts</span>
            <span style="background:{TEAL};color:white;padding:4px 10px;border-radius:6px;font-size:11px;font-weight:700;min-width:50px;text-align:center">{row['opp_score']} {t('pipeline_opp_score', lang)}</span>
          </div>
        </div>""", unsafe_allow_html=True)

    # Download Excel button
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    st.markdown(f'<div style="font-size:11px;font-weight:700;text-transform:uppercase;color:#94A3B8;letter-spacing:0.1em;margin-bottom:10px">EXPORT</div>', unsafe_allow_html=True)
    if st.button(t("pipeline_download", lang), use_container_width=True, key="pipe_dl_btn"):
        try:
            excel_bytes = export_visit_notes_to_excel(
                lang=lang,
                df_rest=df_rest,
                df_rev=df_rev,
                df_ranks_all=df_ranks_all,
                compute_scores_fn=compute_dimension_scores,
                find_col_fn=find_col,
            )
            st.download_button(
                "⬇️ Download",
                data=excel_bytes,
                file_name=f"Vertriebsreporting_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="pipe_dl_btn2",
            )
        except Exception as e:
            st.error(f"Export error: {e}")

# ============================================================
# SILENT WINNERS
# ============================================================
elif st.session_state.active_page == "silent_winners":
    st.markdown(f"""
    <div style="background:{WHITE};border-radius:12px;padding:16px 20px;margin-bottom:20px;border:1px solid {BORDER};box-shadow:0 1px 3px rgba(0,0,0,0.07)">
      <div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#94A3B8;margin-bottom:4px">🌟 {t('silent_winners_title', lang)}</div>
      <div style="font-size:16px;font-weight:700;color:{NAVY}">{t('silent_winners_subtitle', lang)}</div>
      <div style="font-size:11px;color:{MUTED};margin-top:2px">{t('silent_winners_desc', lang)}</div>
    </div>""", unsafe_allow_html=True)

    if _city and _city != "All Cities" and "district" in df_rest.columns:
        city_rest = df_rest[df_rest["district"] == _city]
    else:
        city_rest = df_rest

    silent_winners_city = identify_silent_winners(city_rest)
    if silent_winners_city:
        sw_data = []
        for sw in silent_winners_city[:15]:
            try:
                sw_row    = df_rest[df_rest["name"] == sw].iloc[0]
                sw_scores = compute_dimension_scores(sw, df_rest, df_rev)
                sw_rank   = int(df_ranks_all[df_ranks_all["name"] == sw]["rank"].values[0])
                sw_data.append({
                    t("table_restaurant", lang): sw,
                    t("table_rating", lang):     f"{float(sw_row.get('rating_n', 0)):.1f}⭐",
                    t("table_reviews", lang):    f"{int(sw_row.get('rev_count_n', 0)):,}",
                    t("table_response", lang):   f"{float(sw_row.get('res_rate', 0))*100:.0f}%",
                    t("table_sentiment", lang):  f"{sw_scores['Intelligence']:.0f}%",
                    t("table_rank", lang):        f"#{sw_rank}",
                })
            except Exception as e:
                logger.warning(f"Error processing silent winner {sw}: {e}")
                continue
        if sw_data:
            st.dataframe(pd.DataFrame(sw_data), use_container_width=True, hide_index=True)
            st.markdown(f'<p style="font-size:11px;color:{MUTED};margin-top:8px;">💡 These restaurants have high ratings but low engagement. Focus on <strong>response rate automation</strong> as the opening pitch.</p>', unsafe_allow_html=True)
    else:
        st.info(t("silent_winners_criteria", lang).format(_city))

# FOOTER
st.markdown(f"""
<div style="background:linear-gradient(135deg, {NAVY}12, {TEAL}08);border-top:2px solid {TEAL};border-radius:12px;padding:24px 20px;margin-top:32px;text-align:center">
  <div style="margin-bottom:12px">
    <span style="font-size:28px;margin-right:12px">🍽️</span>
    <span style="font-size:16px;font-weight:800;color:{NAVY};letter-spacing:0.5px">Intelligence Engine v2.0</span>
  </div>
  <p style="font-size:12px;color:{MUTED};margin:8px 0;line-height:1.6">
    <strong style="color:{NAVY}">{len(df_rest)} {t("establishments", lang)}</strong> &nbsp;·&nbsp;
    <strong style="color:{NAVY}">{total} {t("ranked_count", lang)}</strong> &nbsp;·&nbsp;
    <strong style="color:{NAVY}">{t("realtime_analytics", lang)}</strong>
  </p>
  <div style="font-size:10px;color:#94A3B8;margin-top:10px;padding-top:12px;border-top:1px solid {BORDER}">
    <span style="display:inline-block;margin:0 8px"> {t("powered_by_footer", lang)}</span>
  </div>
</div>
""", unsafe_allow_html=True)