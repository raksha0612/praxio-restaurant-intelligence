"""
excel_exporter.py — Restaurant Intelligence Platform
=====================================================
Exports visit notes from the DATABASE (SQLite or PostgreSQL) — NOT from JSON files.

Called by app.py via export_visit_notes_to_excel().
Produces a multi-sheet workbook matching the Vertriebsreporting_Kundenbesuche template.

Sheets:
  1. Visit Notes  — one row per visit, all fields, matches Excel template columns
  2. Pipeline     — all restaurants ranked by score + contacted status
  3. Pipeline KPIs — outcome/interest summary stats
  4. Action Items — upcoming follow-ups sorted by urgency
  5. Images       — full-size embedded attachments (if any)
"""

import base64
import io
import json
import logging
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Colour palette ───────────────────────────────────────────────────────────
TEAL       = "0EA5E9"
TEAL2      = "14B8A6"
NAVY       = "0F172A"
NAVY2      = "1E293B"
GRAY       = "64748B"
LIGHT_BLUE = "F0F9FF"
ALT_ROW    = "F8FAFC"
BORDER_CLR = "E2E8F0"
GREEN_BG   = "DCFCE7"; GREEN_FG = "166534"
RED_BG     = "FEE2E2"; RED_FG   = "991B1B"
AMBER_BG   = "FEF3C7"; AMBER_FG = "92400E"


# ── Style helpers ────────────────────────────────────────────────────────────

def _fill(clr: str) -> PatternFill:
    return PatternFill(start_color=clr, end_color=clr, fill_type="solid")


def _border(clr: str = BORDER_CLR) -> Border:
    s = Side(style="thin", color=clr)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg: str = TEAL) -> None:
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()


def _cell(ws, row: int, col: int, val="", bg: str = "FFFFFF", fg: str = NAVY,
          bold: bool = False, sz: int = 9, h: str = "left", v: str = "center",
          wrap: bool = False) -> object:
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=sz, color=fg, name="Calibri")
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    c.border = _border()
    return c


def _merge(ws, row: int, c1: int, c2: int, val="", bg: str = "FFFFFF",
           fg: str = NAVY, bold: bool = False, sz: int = 9,
           h: str = "left", v: str = "center", wrap: bool = False) -> object:
    ws.merge_cells(
        f"{get_column_letter(c1)}{row}:{get_column_letter(c2)}{row}"
    )
    return _cell(ws, row, c1, val, bg, fg, bold, sz, h, v, wrap)


def _section_hdr(ws, row: int, cols: int, text: str, bg: str = NAVY) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = _fill(bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26


def _kw(date_str: str) -> str:
    """Return 'KW XX' from a date string, or empty string."""
    if not date_str:
        return ""
    try:
        return f"KW {datetime.strptime(str(date_str)[:10], '%Y-%m-%d').isocalendar()[1]}"
    except Exception:
        return ""


def _safe_date(date_str: str) -> datetime:
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(date_str)[:10], fmt)
        except Exception:
            pass
    return datetime(1900, 1, 1)


# ── Main entry point ─────────────────────────────────────────────────────────

def export_from_db(all_calls: list) -> bytes:
    """
    Thin wrapper called by app.py: takes a pre-fetched list of call-note dicts
    (from get_all_call_notes_for_export) and builds the Excel workbook.

    Normalises field names so the sheet builders get what they expect:
    _restaurant_name, visit_date, images list, etc.
    """
    normalised = []
    for call in all_calls:
        c = dict(call)
        rid = c.get("restaurant_id", "")
        c.setdefault("_restaurant_id",   rid)
        c.setdefault("_restaurant_name", rid.replace("_", " ").title())
        c.setdefault("visit_date", c.get("call_date", ""))
        # images may already be parsed by get_all_call_notes_for_export
        if "images" not in c:
            raw = c.get("image_data", "") or ""
            try:
                c["images"] = json.loads(raw) if raw.strip().startswith("[") else []
            except Exception:
                c["images"] = []
        normalised.append(c)

    normalised.sort(key=lambda x: _safe_date(x.get("visit_date", "")), reverse=True)

    wb = Workbook()
    wb.remove(wb.active)
    _sheet_visit_notes(wb, normalised, "EN")
    _sheet_kpis(wb, normalised, "EN")
    _sheet_action_items(wb, normalised, "EN")
    _sheet_images(wb, normalised, "EN")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_visit_notes_to_excel(
    lang: str = "EN",
    df_rest=None,
    df_rev=None,
    df_ranks_all=None,
    compute_scores_fn=None,
    find_col_fn=None,
) -> bytes:
    """
    Build the master Excel workbook from the DATABASE.

    Parameters
    ----------
    lang            : "EN" or "DE"
    df_rest         : restaurants DataFrame (for Pipeline sheet)
    df_rev          : reviews DataFrame (for Pipeline sheet)
    df_ranks_all    : pre-computed rank DataFrame
    compute_scores_fn : callable(name, df_rest, df_rev) → scores dict
    find_col_fn     : callable(df, candidates) → col name or None
    """
    # Import DB functions here so the module can be imported without the full app context
    from database import get_call_notes, get_all_restaurants_with_notes

    all_rids = get_all_restaurants_with_notes()

    # Build flat list of all visits from DB
    all_visits = []
    for rid in sorted(all_rids):
        notes = get_call_notes(rid)
        display_name = rid.replace("_", " ").title()
        for note in notes:
            note["_restaurant_id"]   = rid
            note["_restaurant_name"] = display_name
            all_visits.append(note)

    # Sort newest first
    all_visits.sort(key=lambda x: _safe_date(
        x.get("visit_date") or x.get("call_date", "")
    ), reverse=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Build sheets
    _sheet_visit_notes(wb, all_visits, lang)
    _sheet_pipeline(wb, all_visits, lang, df_rest, df_rev, df_ranks_all,
                    compute_scores_fn, find_col_fn)
    _sheet_kpis(wb, all_visits, lang)
    _sheet_action_items(wb, all_visits, lang)
    _sheet_images(wb, all_visits, lang)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Sheet 1: Visit Notes ─────────────────────────────────────────────────────

def _sheet_visit_notes(wb: Workbook, visits: list, lang: str) -> None:
    """
    One row per visit. Columns exactly match Vertriebsreporting_Kundenbesuche.xlsx.
    """
    is_de = lang.upper() == "DE"
    sheet_name = "Besuchsnotizen" if is_de else "Visit Notes"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Column definitions: (header_EN, header_DE, db_key, width)
    COLS = [
        ("KW",                                  "KW",                                   "_kw",              7),
        ("Restaurant",                          "Kunde",                                "_restaurant_name", 28),
        ("Visit Date",                          "Besuchsdatum",                         "visit_date",       14),
        ("Time",                                "Uhrzeit",                              "visit_time",       10),
        ("City",                                "Stadt",                                "city",             14),
        ("District",                            "Stadtteil",                            "district",         14),
        ("Price Class",                         "Preisklasse",                          "price_class",      12),
        ("Size",                                "Größe",                                "size",             12),
        ("Contact Person",                      "Gesprächspartner",                     "contact_name",     24),
        ("Atmosphere on Site",                  "Stimmung vor Ort",                     "atmosphere",       26),
        ("Visit Duration",                      "Dauer des Gesprächs",                  "visit_duration",   14),
        ("Pre-Check Needs",                     "Bedarf Pre-Check",                     "pre_check_needs",  30),
        ("Potential Estimate (1-10)",            "Einschätzung des Potenzials (1-10)",   "potential_score",  14),
        ("Interest Level (1-5)",                "Interessensstufe (1-5)",               "interest_level",   14),
        ("Products Shown",                      "Gezeigte Produkte",                    "_products",        30),
        ("Outcome",                             "Ergebnis",                             "visit_outcome",    16),
        ("Next Steps / Follow-up Plan",         "Geplantes Follow up",                  "next_steps",       30),
        ("Follow-up Date",                      "Nachverfolgungsdatum",                 "followup_date",    14),
        ("Detailed Notes",                      "Ausführliche Notizen",                 "notes",            50),
        ("Self-Reflection (for Kevin)",         "Selbstreflexion (Basis für Dialog mit Kevin)", "self_reflection", 50),
        ("Main Objection",                      "Haupteinwand",                         "main_objection",   24),
        ("Budget Range",                        "Budget",                               "budget_range",     16),
        ("Confidence Level (Close %)",          "Abschluss-Sicherheit %",               "confidence",       14),
        ("Decision Timeline",                   "Entscheidungsfrist",                   "decision_timeline",20),
        ("Competitor Tools Mentioned",          "Konkurrenztools",                      "competitor_tools", 22),
    ]

    # Title row
    n_cols = len(COLS)
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    title_text = (
        f"  🍽️  Praxiotech — {'Besuchsprotokoll' if is_de else 'Visit Log'}   ·   "
        f"Exported {datetime.now().strftime('%d %b %Y  %H:%M')}   ·   "
        f"{len(visits)} {'Besuch(e)' if is_de else 'visit(s)'}"
    )
    t = ws.cell(row=1, column=1, value=title_text)
    t.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t.fill = _fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    # Header row
    for ci, (hdr_en, hdr_de, _, col_w) in enumerate(COLS, 1):
        hdr_text = hdr_de if is_de else hdr_en
        cell = ws.cell(row=2, column=ci, value=hdr_text)
        _hdr(cell, bg=TEAL)
        ws.column_dimensions[get_column_letter(ci)].width = col_w
    ws.row_dimensions[2].height = 30

    # Data rows
    outcome_colors = {
        "won":              (GREEN_BG, GREEN_FG),
        "gewonnen":         (GREEN_BG, GREEN_FG),
        "lost":             (RED_BG,   RED_FG),
        "verloren":         (RED_BG,   RED_FG),
        "demo scheduled":   (LIGHT_BLUE, "0369A1"),
        "demo vereinbart":  (LIGHT_BLUE, "0369A1"),
        "proposal sent":    ("EDE9FE", "5B21B6"),
        "angebot gesendet": ("EDE9FE", "5B21B6"),
        "interested":       ("FEF9C3", "854D0E"),
        "interessiert":     ("FEF9C3", "854D0E"),
    }

    for row_i, visit in enumerate(visits, 3):
        is_alt = row_i % 2 == 0
        base_bg = ALT_ROW if is_alt else "FFFFFF"
        outcome_raw = (visit.get("visit_outcome") or visit.get("outcome") or "Pending").strip()
        out_bg, out_fg = outcome_colors.get(outcome_raw.lower(), (AMBER_BG, AMBER_FG))

        products_str = ", ".join(
            visit.get("products_discussed") or []
        )

        for ci, (_, _, db_key, _) in enumerate(COLS, 1):
            if db_key == "_kw":
                val = _kw(visit.get("visit_date") or visit.get("call_date", ""))
            elif db_key == "_restaurant_name":
                val = visit.get("_restaurant_name", "")
            elif db_key == "_products":
                val = products_str
            elif db_key == "visit_outcome":
                val = outcome_raw
            else:
                val = visit.get(db_key, "") or ""

            cell = ws.cell(row=row_i, column=ci, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            # Colour special columns
            if db_key == "visit_outcome":
                cell.fill = _fill(out_bg)
                cell.font = Font(bold=True, size=9, color=out_fg, name="Calibri")
            elif db_key in ("notes", "self_reflection"):
                cell.fill = _fill("FFFBEB" if not is_alt else "FEF9C3")
            elif db_key == "potential_score":
                try:
                    pv = int(val)
                    bg = GREEN_BG if pv >= 8 else (AMBER_BG if pv >= 5 else RED_BG)
                    fg = GREEN_FG if pv >= 8 else (AMBER_FG if pv >= 5 else RED_FG)
                    cell.fill = _fill(bg)
                    cell.font = Font(bold=True, size=9, color=fg, name="Calibri")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except Exception:
                    cell.fill = _fill(base_bg)
            elif db_key == "interest_level":
                try:
                    iv = int(val)
                    stars = "★" * iv + "☆" * (5 - iv)
                    cell.value = f"{stars}  {iv}/5"
                    bg = GREEN_BG if iv >= 4 else (AMBER_BG if iv >= 2 else RED_BG)
                    fg = GREEN_FG if iv >= 4 else (AMBER_FG if iv >= 2 else RED_FG)
                    cell.fill = _fill(bg)
                    cell.font = Font(bold=True, size=9, color=fg, name="Calibri")
                except Exception:
                    cell.fill = _fill(base_bg)
            else:
                cell.fill = _fill(base_bg)

        # Row height — taller for note columns
        ws.row_dimensions[row_i].height = 60

    ws.freeze_panes = "A3"

    # Empty state
    if not visits:
        ws.merge_cells(f"A3:{get_column_letter(n_cols)}3")
        empty = ws.cell(row=3, column=1,
                        value="No visits logged yet. Use the Visit Log tab to add your first entry.")
        empty.font = Font(italic=True, size=11, color=GRAY, name="Calibri")
        empty.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 40


# ── Sheet 2: Pipeline ────────────────────────────────────────────────────────

def _sheet_pipeline(wb, visits, lang, df_rest, df_rev, df_ranks_all,
                    compute_scores_fn, find_col_fn):
    is_de = lang.upper() == "DE"
    ws = wb.create_sheet("Vertriebspipeline" if is_de else "Pipeline")
    ws.sheet_view.showGridLines = False

    if df_ranks_all is None or compute_scores_fn is None:
        # Fallback: simple restaurant list from visit notes
        from database import get_all_restaurants_with_notes
        _section_hdr(ws, 1, 6, "  📊  Pipeline — Visit Summary")
        hdrs = ["Restaurant", "Total Visits", "Last Visit", "Avg Interest", "Last Outcome", "Follow-up Date"]
        if is_de:
            hdrs = ["Kunde", "Besuche Gesamt", "Letzter Besuch", "Ø Interesse", "Letztes Ergebnis", "Follow-up Datum"]
        for ci, h in enumerate(hdrs, 1):
            _hdr(ws.cell(row=2, column=ci), bg=TEAL)
            ws.cell(row=2, column=ci).value = h
            ws.column_dimensions[get_column_letter(ci)].width = [28,14,14,14,20,16][ci-1]
        ws.row_dimensions[2].height = 22

        from database import get_all_restaurants_with_notes, get_call_notes
        all_rids = get_all_restaurants_with_notes()
        row = 3
        for rid in sorted(all_rids):
            notes = get_call_notes(rid)
            if not notes:
                continue
            last = notes[-1]
            avg_int = sum(n.get("interest_level") or 0 for n in notes) / len(notes)
            bg = ALT_ROW if row % 2 == 0 else "FFFFFF"
            vals = [
                rid.replace("_"," ").title(),
                len(notes),
                last.get("visit_date") or last.get("call_date",""),
                f"{avg_int:.1f}/5",
                last.get("visit_outcome") or last.get("outcome",""),
                last.get("followup_date",""),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = Font(size=9, name="Calibri")
                cell.fill = _fill(bg)
                cell.border = _border()
                cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
        ws.freeze_panes = "A3"
        return

    from database import get_all_restaurants_with_notes, get_call_notes
    all_rids_with_notes = set(get_all_restaurants_with_notes())

    _section_hdr(ws, 1, 10,
                 "  📊  Vertriebspipeline — Opportunity Ranking" if is_de
                 else "  📊  Sales Pipeline — Opportunity Ranking")
    hdrs_en = ["Rank","Restaurant","City","Score","Rating","Reviews","Response %","Tag","Contacted","Visits"]
    hdrs_de = ["Rang","Kunde","Stadt","Score","Bewertung","Rezensionen","Antwortquote","Tag","Kontaktiert","Besuche"]
    hdrs = hdrs_de if is_de else hdrs_en
    col_widths = [7, 30, 14, 10, 10, 12, 14, 20, 12, 10]

    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        _hdr(cell, bg=TEAL)
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22

    row = 3
    for _, rank_row in df_ranks_all.iterrows():
        name = rank_row["name"]
        try:
            res_row = df_rest[df_rest["name"] == name].iloc[0]
            s = compute_scores_fn(name, df_rest, df_rev)
        except Exception:
            continue
        rid = name.lower().replace(" ","_").replace("-","_")[:40]
        is_contacted = rid in all_rids_with_notes
        n_visits = len(get_call_notes(rid))

        city_val = ""
        if find_col_fn:
            dc = find_col_fn(df_rest, ["district"])
            if dc:
                city_val = str(res_row.get(dc, ""))

        is_alt = row % 2 == 0
        base_bg = ALT_ROW if is_alt else "FFFFFF"

        vals = [
            int(rank_row["rank"]),
            name,
            city_val,
            round(float(s.get("Composite", 0)), 1),
            round(float(res_row.get("rating_n", 0)), 1),
            int(res_row.get("rev_count_n", 0)),
            f"{float(res_row.get('res_rate', 0))*100:.0f}%",
            "✅ Contacted" if is_contacted else "○ Not contacted",
            "Yes" if is_contacted else "No",
            n_visits,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = Font(size=9, name="Calibri")
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if ci != 2 else "left",
                                       vertical="center")
            if ci == 8:
                if is_contacted:
                    cell.fill = _fill(GREEN_BG)
                    cell.font = Font(bold=True, size=9, color=GREEN_FG, name="Calibri")
                else:
                    cell.fill = _fill(ALT_ROW)
                    cell.font = Font(size=9, color=GRAY, name="Calibri")
            else:
                cell.fill = _fill(base_bg)
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A3"


# ── Sheet 3: KPIs ────────────────────────────────────────────────────────────

def _sheet_kpis(wb, visits, lang):
    is_de = lang.upper() == "DE"
    ws = wb.create_sheet("KPIs")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16

    _section_hdr(ws, 1, 3,
                 "  📈  Pipeline KPIs & Auswertung" if is_de
                 else "  📈  Pipeline KPIs & Summary")
    ws.merge_cells("A2:C2")
    sub = ws.cell(row=2, column=1,
                  value=f"  Exported {datetime.now().strftime('%d %b %Y  %H:%M')}  ·  {len(visits)} visits")
    sub.font = Font(size=8, color="94A3B8", name="Calibri")
    sub.fill = _fill(NAVY2)
    sub.alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 14

    # Outcome breakdown
    outcomes = {}
    for v in visits:
        o = (v.get("visit_outcome") or v.get("outcome") or "Pending").strip()
        outcomes[o] = outcomes.get(o, 0) + 1

    interest_vals = [v.get("interest_level") or 0 for v in visits]
    avg_interest = sum(interest_vals) / len(interest_vals) if interest_vals else 0

    products_count: dict = {}
    for v in visits:
        for p in (v.get("products_discussed") or []):
            products_count[p] = products_count.get(p, 0) + 1

    row = 4
    kpi_label = "Kennzahl" if is_de else "KPI"
    val_label  = "Wert" if is_de else "Value"
    for ci, h in enumerate([kpi_label, val_label], 1):
        cell = ws.cell(row=row, column=ci, value=h)
        _hdr(cell, bg=NAVY)
    ws.row_dimensions[row].height = 22
    row += 1

    kpis_en = [
        ("Total Visits Logged",           len(visits),                     "FFFFFF", NAVY),
        ("Restaurants Contacted",          len(set(v["_restaurant_id"] for v in visits)), "FFFFFF", NAVY),
        ("Average Interest Level",         f"{avg_interest:.1f} / 5",       "FFFFFF", NAVY),
        ("High Interest (4-5 ★)",          sum(1 for x in interest_vals if x >= 4), GREEN_BG, GREEN_FG),
        ("Medium Interest (2-3 ★)",        sum(1 for x in interest_vals if 2 <= x < 4), AMBER_BG, AMBER_FG),
        ("Low Interest (1 ★)",             sum(1 for x in interest_vals if x == 1), RED_BG, RED_FG),
        ("Visits with Images",             sum(1 for v in visits if v.get("images")), LIGHT_BLUE, "0369A1"),
    ]
    kpis_de = [
        ("Besuche gesamt",                 len(visits),                     "FFFFFF", NAVY),
        ("Kontaktierte Restaurants",       len(set(v["_restaurant_id"] for v in visits)), "FFFFFF", NAVY),
        ("Ø Interessensstufe",             f"{avg_interest:.1f} / 5",       "FFFFFF", NAVY),
        ("Hohes Interesse (4-5 ★)",        sum(1 for x in interest_vals if x >= 4), GREEN_BG, GREEN_FG),
        ("Mittleres Interesse (2-3 ★)",    sum(1 for x in interest_vals if 2 <= x < 4), AMBER_BG, AMBER_FG),
        ("Geringes Interesse (1 ★)",       sum(1 for x in interest_vals if x == 1), RED_BG, RED_FG),
        ("Besuche mit Bildern",            sum(1 for v in visits if v.get("images")), LIGHT_BLUE, "0369A1"),
    ]
    kpis = kpis_de if is_de else kpis_en

    for label, value, bg, fg in kpis:
        for ci, val in enumerate([label, value], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = _fill(bg)
            cell.font = Font(bold=(ci == 1), size=10, color=fg, name="Calibri")
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

    # Outcome breakdown
    row += 1
    _section_hdr(ws, row, 3,
                 "  🏁  Ergebnisverteilung" if is_de else "  🏁  Outcome Breakdown",
                 bg=TEAL)
    row += 1
    for o, cnt in sorted(outcomes.items(), key=lambda x: x[1], reverse=True):
        ok = o.lower()
        if "won" in ok or "gewonnen" in ok:
            bg, fg = GREEN_BG, GREEN_FG
        elif "lost" in ok or "verloren" in ok:
            bg, fg = RED_BG, RED_FG
        else:
            bg, fg = AMBER_BG, AMBER_FG
        pct = f"{cnt/max(len(visits),1)*100:.1f}%"
        for ci, val in enumerate([o, cnt, pct], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = _fill(bg)
            cell.font = Font(bold=(ci == 1), size=10, color=fg, name="Calibri")
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

    # Product frequency
    if products_count:
        row += 1
        _section_hdr(ws, row, 3,
                     "  📦  Gezeigte Produkte" if is_de else "  📦  Products Shown Frequency",
                     bg=TEAL2)
        row += 1
        for p, cnt in sorted(products_count.items(), key=lambda x: x[1], reverse=True):
            for ci, val in enumerate([p, cnt], 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.fill = _fill(LIGHT_BLUE if row % 2 == 0 else "FFFFFF")
                cell.font = Font(bold=(ci == 2), size=9,
                                 color=TEAL if ci == 2 else NAVY, name="Calibri")
                cell.alignment = Alignment(horizontal="left" if ci == 1 else "center",
                                           vertical="center")
                cell.border = _border()
            ws.row_dimensions[row].height = 16
            row += 1


# ── Sheet 4: Action Items ────────────────────────────────────────────────────

def _sheet_action_items(wb, visits, lang):
    is_de = lang.upper() == "DE"
    ws = wb.create_sheet("Aktionen" if is_de else "Action Items")
    ws.sheet_view.showGridLines = False

    _section_hdr(ws, 1, 6,
                 "  ⚡  Anstehende Follow-ups — nach Dringlichkeit" if is_de
                 else "  ⚡  Upcoming Follow-ups — sorted by urgency")

    hdrs_en = ["Restaurant", "Last Visit", "Contact", "Next Steps", "Days Until Due", "Status"]
    hdrs_de = ["Kunde",      "Letzter Besuch", "Gesprächspartner", "Geplantes Follow up", "Tage bis Fälligkeit", "Status"]
    hdrs    = hdrs_de if is_de else hdrs_en
    col_w   = [28, 14, 22, 34, 16, 14]

    for ci, (h, w) in enumerate(zip(hdrs, col_w), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        _hdr(cell, bg=TEAL)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    # Group by restaurant, keep last visit per restaurant
    by_rid: dict = {}
    for v in visits:
        rid = v["_restaurant_id"]
        existing = by_rid.get(rid)
        if existing is None or _safe_date(v.get("visit_date") or "") > _safe_date(existing.get("visit_date") or ""):
            by_rid[rid] = v

    today = datetime.now()
    rows_data = []
    for rid, last in by_rid.items():
        fu_date_str = last.get("followup_date", "")
        if fu_date_str:
            try:
                fu_date = _safe_date(fu_date_str)
                days = (fu_date - today).days
            except Exception:
                days = 7
        else:
            visit_dt = _safe_date(last.get("visit_date") or last.get("call_date", ""))
            days = (visit_dt + timedelta(days=7) - today).days

        status = ("OVERDUE" if days < 0 else ("URGENT" if days <= 2 else "PENDING"))
        if is_de:
            status = ("ÜBERFÄLLIG" if days < 0 else ("DRINGEND" if days <= 2 else "AUSSTEHEND"))

        rows_data.append((days, last["_restaurant_name"], last, status))

    # Sort: overdue first, then by days ascending
    rows_data.sort(key=lambda x: x[0])

    row = 3
    for days, rest_name, last, status in rows_data:
        if "OVERDUE" in status or "ÜBERFÄLLIG" in status:
            sbg, sfg = RED_BG, RED_FG
        elif "URGENT" in status or "DRINGEND" in status:
            sbg, sfg = AMBER_BG, AMBER_FG
        else:
            sbg, sfg = GREEN_BG, GREEN_FG

        base_bg = ALT_ROW if row % 2 == 0 else "FFFFFF"
        vals = [
            rest_name,
            last.get("visit_date") or last.get("call_date", ""),
            last.get("contact_name", ""),
            (last.get("next_steps", "") or "")[:80],
            max(days, 0),
            status,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="center",
                                       horizontal="left" if ci in (1, 3, 4) else "center")
            if ci == 6:
                cell.fill = _fill(sbg)
                cell.font = Font(bold=True, size=9, color=sfg, name="Calibri")
            else:
                cell.fill = _fill(base_bg)
                cell.font = Font(size=9, name="Calibri")
        ws.row_dimensions[row].height = 18
        row += 1

    ws.freeze_panes = "A3"


# ── Sheet 5: Images ──────────────────────────────────────────────────────────

def _sheet_images(wb, visits, lang):
    is_de = lang.upper() == "DE"
    try:
        from openpyxl.drawing.image import Image as XLImage
        can_embed = True
    except ImportError:
        can_embed = False

    ws = wb.create_sheet("📸 Bilder" if is_de else "📸 Images")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 2

    ws.merge_cells("A1:C1")
    ws["A1"] = ("  📸  Bildergalerie — Vollbild-Anhänge" if is_de
                else "  📸  Image Gallery — Full Size Attachments")
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:C2")
    ws["A2"] = "  Full resolution images from visit notes."
    ws["A2"].font = Font(size=8, color="94A3B8", italic=True, name="Calibri")
    ws["A2"].fill = _fill(NAVY2)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    g_row = 4
    any_images = False

    for visit in visits:
        images = visit.get("images") or []
        if not images:
            continue
        any_images = True
        rest_name  = visit.get("_restaurant_name", "")
        visit_date = visit.get("visit_date") or visit.get("call_date", "")

        for img_idx, img_data in enumerate(images):
            filename = img_data.get("filename", f"image_{img_idx+1}")

            ws.merge_cells(f"A{g_row}:C{g_row}")
            hdr = ws.cell(row=g_row, column=1,
                          value=f"  📍  {rest_name}  ·  {visit_date}  ·  {filename}")
            hdr.font = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            hdr.fill = _fill(TEAL)
            hdr.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[g_row].height = 22
            g_row += 1

            if can_embed and img_data.get("data"):
                try:
                    raw = base64.b64decode(img_data["data"])
                    xl = XLImage(io.BytesIO(raw))
                    ratio = min(580 / max(xl.width, 1), 440 / max(xl.height, 1))
                    xl.width  = int(xl.width * ratio)
                    xl.height = int(xl.height * ratio)
                    ws.row_dimensions[g_row].height = int(xl.height * 0.75) + 8
                    ws.add_image(xl, f"B{g_row}")
                except Exception as e:
                    logger.warning("Image embed error: %s", e)
                    ws.cell(row=g_row, column=2,
                            value=f"[Could not embed: {filename}]").font = Font(
                        italic=True, size=9, color=GRAY, name="Calibri")
                    ws.row_dimensions[g_row].height = 20
            else:
                ws.cell(row=g_row, column=2,
                        value=f"[{filename}]").font = Font(
                    italic=True, size=9, color=GRAY, name="Calibri")
                ws.row_dimensions[g_row].height = 20

            g_row += 2

    if not any_images:
        ws.merge_cells("A4:C4")
        ws["A4"] = ("Noch keine Bilder angehängt." if is_de
                    else "No images attached yet.")
        ws["A4"].font = Font(italic=True, size=10, color=GRAY, name="Calibri")
        ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 30


# ── Backward-compat shim so old calls don't crash ───────────────────────────

def export_call_notes_to_excel(call_notes_dir=None, lang: str = "EN", **kwargs) -> bytes:
    """
    Legacy entry point kept for backward compatibility.
    Ignores call_notes_dir — reads from DB instead.
    """
    logger.warning(
        "export_call_notes_to_excel() called with call_notes_dir — "
        "ignoring JSON path, reading from database instead."
    )
    return export_visit_notes_to_excel(lang=lang, **kwargs)